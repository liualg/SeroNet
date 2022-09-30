"""
This script is still in Beta testing. Currently it will take a SeroNet Registry template and 
convert it into a ImmPort template. 

*** Things to do before runnning the script:
- install the requirements 

- Update the field: PMID
    - This will be the PMID associated with the SeroNet Registry template. It should be located
    in the Box folder with this naming scheme "registry_XXXXXX"

- update the field: box_base
    - This will be the file path to the "SeroNet Public Data" within Box. This is something that 
    must be integrated into your file directory system 

A.Liu - Mar 17, 2022
"""

import pandas as pd 
import numpy as np
import os, shutil
import inspect

from tqdm import tqdm

from dataclasses import dataclass, field
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

## Importing Functions and Dataclass
import seronetDataclass_v1 as seroClass
import seronetFunctions as seroFxn

warnings = 0
warn = 0


################## FILLOUT THIS SECTION  ###################
PMID = "33184236"
# PMID = "34270939"

# Registry
sheet_name = 'SeroNet Registry Template'
map_sheet = 'Registry Definitions'


# Note: box needs to be integrated into your file system

box_base = "~/Library/CloudStorage/Box-Box/SeroNet Public Data"

##################          CODE          ###################


#File Paths
BASE_DIR = seroFxn.get_box_dir(box_base, PMID)


file = "registry_" + PMID
df_path = os.path.join(BASE_DIR,'templated_data', file + ".xlsm")
# df_path = "./registry_33184236.xlsm"

# dictionary info
registryToImmportDict_file = os.path.join("dictionary", "registryToBasic.csv")

registryToImmportDict = pd.read_csv(registryToImmportDict_file,
                                    header=None, 
                                    index_col=0, 
                                    squeeze=True).to_dict()

# ImmPort Templates (link to web?)
PATH_basic_stdy_template = os.path.join("template", "basic_study_design.xlsx")
PATH_protocols = os.path.join("template", "protocols.xlsx")
PATH_experiments = os.path.join("template", "experiments.xlsx")
PATH_reagent = os.path.join("template", "reagents.Other.xlsx")
PATH_assessment = os.path.join("template", "assessments.xlsx")
PATH_subject_human = os.path.join("template", "subjectHuman.xlsx")
PATH_subject_organism = os.path.join("template", "subjectAnimals.xlsx")


# Automate output... 
OUT_DIR = os.path.join(BASE_DIR, 'ImmPort_templates') 
# OUT_DIR = './33184236_test/'
PATH_pmid_basic_stdy_template = f'PMID{PMID}_study.xlsx'

BASIC_STUDY_TEMPLATE = f'PMID{PMID}_basic'
EXP_TEMPLATE = f'PMID{PMID}_experiments'
PROTOCOL_TEMPLATE = f'PMID{PMID}_protocol'
REAGENT_TEMPLATE = f'PMID{PMID}_reagent'
ASSESSMENT_TEMPLATE =  f'PMID{PMID}_assessment'
SUBJ_HUMAN_TEMPLATE =  f'PMID{PMID}_subject_human'
SUBJ_ORGANISM_TEMPLATE =  f'PMID{PMID}_subject_organism'

# Make Dir if it does not exist
try:
    os.mkdir(OUT_DIR)
    print(f'Creating output directory - {OUT_DIR}')
except FileExistsError:
    print('Will not create directory - already exists')
    pass

# import file
book = load_workbook(df_path)
registry = book[sheet_name]
registry.delete_cols(1)


# Class names
class_names = ['study_pubmed', 'study', 'study_personnel', 'study_file',
               'study_link', 'study_categorization', 'study_design',
               'protocol', 'condition_or_disease', 'Intervention Agent',
               'study_details', 'inclusion_exclusion',
               'Subjects Type: human', 'Subjects Type: model organism', 'planned_visit',
               'Study Experiments', 'Reagent per Experiment-Assay Type', 
               'Results'
              ]

sp = seroFxn.get_sections(registry, class_names)

### Main Loop

# Looping through each section in the Registy template
for section_number in range(len(sp)-1):
    temp_wb = Workbook()
    temp_ws = temp_wb.active

    #making a temp workbook to store each section. This will be turned into df
    for i in registry.iter_rows(values_only = True,
                                min_row = sp[section_number]+1,
                                max_row = sp[section_number+1]-1):
        temp_ws.append(i)
        
    max_row = temp_ws.max_row
    max_col = temp_ws.max_column
    seroFxn.remove_excess(temp_ws)
    
    df = pd.DataFrame(temp_ws.values)
    sub_section = registry.cell(row=sp[section_number], column = 1).value

    if sub_section == 'study':
        df = seroFxn.edit_df(df)
        
        STUDY = seroClass.study(
            df['Study Identifier'][1],
            df['Study Name'][1],
            df['Publication Title'][1],
            df['Study Objective'][1],
            df['Study Description'][1],
            df['Primary Institution Name'][1]
            )
    
    elif sub_section == 'study_pubmed':
        df = seroFxn.edit_df(df)
        
        STUDY_PUBMED = seroClass.study_pubmed(
            df['Pubmed ID'][1]
        )
        
    elif sub_section == 'study_personnel':
        df = seroFxn.edit_df(df)

        STUDY_PERSONNEL = seroClass.study_personnel(
            df['Personnel ID'],
            df['Honorific'],
            df['Last Name'],
            df['First Name'],
            df['Suffixes'],
            df['Organization'],
            df['ORCID ID'],
            df['Email'],
            df['Title In Study'],
            df['Role In Study'],
            df['Site Name']
        )


    elif sub_section == 'study_file':
        df = seroFxn.edit_df(df)

        STUDY_FILE = seroClass.study_file(
            df['Study File Name'],
            df['Study File Description'],
            df['Study File Type']  
        )
        
    
    elif sub_section == 'study_link':
        df = seroFxn.edit_df(df)

        STUDY_LINK = seroClass.study_link(
            df['Link Name'],
            df['Value']
        )
             
        
    elif sub_section == 'study_categorization':
        df = seroFxn.edit_df(df)

        STUDY_CATEGORIZATION = seroClass.study_categorization(
            df['Research Focus*'][1],
            df['Study Type'][1],
            df['Keywords'][1],
        )
    
    elif sub_section == 'study_design':
        df = seroFxn.edit_df(df)

        if (df.shape != (2,0)): # checking size. There has to be a better way to do this

            STUDY_DESIGN = seroClass.study_design(
                df['Clinical Study Design'],
                df['in silico Model Type*']
            )
        else:
            study_des = study_design()

    elif sub_section == 'protocol':
        df = seroFxn.edit_df(df)

        PROTOCOLS = seroClass.protocols(
            df['Protocol ID'],
            df['Protocol File Name'],
            df['Protocol Name'],
            df['Protocol Description'],
            df['Protocol Type'],
        )
        
    elif sub_section == 'condition_or_disease':
        df = seroFxn.edit_df(df)

        COD = seroClass.condition_or_disease(
            list(df['Reported Health Condition* '])
        )
        
    elif sub_section == 'Intervention Agent':
        df = seroFxn.edit_df(df)

        INTERVENTION_AGENT = seroClass.Intervention_Agent(
            list(df['SARS-CoV-2 Vaccine Type*'])
        )

    elif sub_section == 'study_details':
        df = seroFxn.edit_df(df)
        STUDY_DETAILS = seroClass.study_details(
            df['Clinical Outcome Measure'][1],
            df['Enrollment Start Date'][1],
            df['Enrollment End Date'][1],
            df['Number of Study Subjects'][1],
            df['Age Unit'][1],
            df['Minimum Age'][1],
            df['Maximum Age'][1]
            )

    elif sub_section == 'inclusion_exclusion':
        df = seroFxn.edit_df(df)

        INCLUSION_EXCLUSION = seroClass.inclusion_exclusion(
            df['Inclusion ID'],
            df['Inclusion Criterion'],
            df['Inclusion Criterion Category']  
        )
        
    elif sub_section == 'Subjects Type: human':
        df = seroFxn.edit_df(df)

        SUBJECT_HUMAN = seroClass.subject_type_human(
            df['Arm ID'],
            df['Arm Name'],
            df['Study Population Description'],
            df['Arm Type'],
            df['Ethnicity*'],
            df['Race*'],
            df['Race Specify'],
            df['Description'],
            df['Sex at Birth*'],
            df['Age Event'],
            df['Subject Phenotype'],
            df['Study Location*'],
            df['Assessment Name'],
            df['Measured Behavioral or Psychological Factor'],
            df['Measured Social Factor'],
            df['SARS-CoV-2 Symptoms'],
            df['Assessment_Clinical  and Demographic Data Provenance'],
            df['Assessment_Demographic Data Types Collected'],
            df['SARS-CoV2 History*'],
            df['SARS-CoV-2 Vaccine Type*'],
            df['COVID-19 Disease Severity*']
            )

    elif sub_section == 'Subjects Type: model organism':
        df = seroFxn.edit_df(df)

        SUBJECT_ORGANISM = seroClass.subject_type_mode_organism(
            df['Arm ID'],
            df['Arm Name'],
            df['Study Population Description'],
            df['Arm Type'],
            df['Species'],
            df['Biosample Types'],
            df['Strain Characteristics'],
            df['Sex at Birth*'],
            df['Age Event'],
            df['Subject Phenotype'],
            df['Study Location*'],
            df['SARS-CoV-2 Symptoms'],
            df['SARS-CoV2 History*'],
            df['SARS-CoV-2 Vaccine Type*'],
            df['COVID-19 Disease Severity*']
            )
    
    elif sub_section == 'planned_visit':
        df = seroFxn.edit_df(df)

        PLANNED_VISIT = seroClass.planned_visit(
            df['Visit ID'],
            df['Visit Name'],
            df['Visit Order Number'],
            df['Visit Min Start Day'],
            df['Visit Max Start Day'],
            df['Visit Start Rule']                                
    )
    
#     elif sub_section == 'Study Experiment Samples':
#         df = seroFxn.edit_df(df)

#         EXPERIMENT_SAMPLES = seroClass.study_experiment_samples(
#             df['Expt Sample User Defined ID'],
#             df['Expt Sample Biospecimen Type'],
#             df['Expt Sample Biospecimen Collection Point']                             
#     )
    
    elif sub_section == 'Study Experiments':
        df = seroFxn.edit_df(df)

        STUDY_EXPERIMENTS = seroClass.study_experiment(
            df['Experiment ID'],
            df['Experiment Name'],
            df['Experiment Assay Type'],
            df['Experiment Results File Name'],
            df['Expt Sample Biospecimen Type'],
            df['Expt Sample Biospecimen Collection Point']
    )
        
    elif sub_section == 'Reagent per Experiment-Assay Type':
        df = seroFxn.edit_df(df)

        REAGENTS = seroClass.reagent_per_experiment(
            df['Reagent ID'],
            df['SARS-CoV-2 Antigen'],
            df['Assay Use'],
            df['Manufacturer'],
            df['Catalog #']
    )  
        
    elif sub_section == 'Results':
        df = seroFxn.edit_df(df)

        RESULTS = seroClass.results(
            df['Results Virus Target'],
            df['Results Antibody Isotype'],
            df['Results Reporting Units'],
            df['Results Reporting Format']
    )  
        
    else:
        print(sub_section, ': does not exist')




# Sending back to ImmPort Template #

ImmPortClassNames = ['study', 'study_categorization', 'study_2_condition_or_disease',
               'arm_or_cohort','study_personnel', 'planned_visit', 'inclusion_exclusion',
               'study_2_protocol', 'study_file', 'study_link', 'study_pubmed'] 

shutil.copy(PATH_basic_stdy_template, PATH_pmid_basic_stdy_template)
basic_stdy_template = load_workbook(PATH_pmid_basic_stdy_template)
bst_ws = basic_stdy_template['basic_study_design.txt']

se = seroFxn.get_sections(bst_ws, ImmPortClassNames)


# Basic Study Template

if len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type) + len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
    
    vaccine_type = list(set([i.split(';')[0] for i in SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type] + [i.split(';')[0] for i in SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type]))
    
elif len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type):
    
    vaccine_type = list(set([i.split(';')[0] for i in SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type]))
                            
elif len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
    
    vaccine_type = list(set([i.split(';')[0] for i in SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type]))
    
else:
    
    vaccine_type = []

if SUBJECT_HUMAN and SUBJECT_ORGANISM:
    print('Both human and model organism are used')
    usr_id = list(SUBJECT_HUMAN.User_Defined_ID) + list(SUBJECT_ORGANISM.User_Defined_ID)
    name = list(SUBJECT_HUMAN.Name) + list(SUBJECT_ORGANISM.Name)
    description = list(SUBJECT_HUMAN.Description) + list(SUBJECT_ORGANISM.Description)
    type_report = list(SUBJECT_HUMAN.Type_Reported) + list(SUBJECT_ORGANISM.Type_Reported)
    
if SUBJECT_HUMAN and not SUBJECT_ORGANISM:
    print('Only human is used')
    usr_id = list(SUBJECT_HUMAN.User_Defined_ID)
    name = list(SUBJECT_HUMAN.Name)
    description = list(SUBJECT_HUMAN.Description)
    type_report = list(SUBJECT_HUMAN.Type_Reported)
    vaccine_type = [i.split(';')[0] for i in list(set(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type))]
    
if SUBJECT_ORGANISM and not SUBJECT_HUMAN:
    print('Only organism is used')
    usr_id = list(SUBJECT_ORGANISM.User_Defined_ID)
    name = list(SUBJECT_ORGANISM.Name)
    description = list(SUBJECT_ORGANISM.Description)
    type_report = list(SUBJECT_ORGANISM.Type_Reported)
    vaccine_type = [i.split(';')[0] for i in list(set(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type))]

AOC = seroClass.arm_or_cohort(
    list(filter(None, usr_id)),
    list(filter(None, name)),
    list(filter(None, description)),
    list(filter(None, type_report))
)

temp_wb = Workbook()
temp_ws = temp_wb.active

#making a temp workbook to store each section. This will be turned into df

# This is the first part of the df
for i in bst_ws.iter_rows(values_only = True,
                          max_row = se[2]+1):
    temp_ws.append(i)
    
# Starting at Arm or Cohort  



AOC = seroClass.arm_or_cohort(
    list(filter(None, usr_id)),
    list(filter(None, name)),
    list(filter(None, description)),
    list(filter(None, type_report))
)

# temp_ws.append(['User Defined ID', 'Name','Description','Type Reported'])
seroFxn.add_df(temp_ws, AOC)

seroFxn.add_df(temp_ws, STUDY_PERSONNEL)
seroFxn.add_df(temp_ws, PLANNED_VISIT)
seroFxn.add_df(temp_ws, INCLUSION_EXCLUSION)

# # Add protocol
temp_ws.append([])
temp_ws.append([PROTOCOLS.ImmPortNAME])
temp_ws.append(['Protocol ID', PROTOCOLS.Protocol_Name[1]])

seroFxn.add_df(temp_ws, STUDY_FILE)
seroFxn.add_df(temp_ws, STUDY_LINK)

seroFxn.add_df(temp_ws, STUDY_PUBMED)

pd.DataFrame(temp_ws.values).to_csv('./temp_mar6.csv', index=None)

#### Filling in Study Info Sections 

registryToImmportDict = pd.read_csv(registryToImmportDict_file, 
                                    header=None, 
                                    index_col=0, 
                                    squeeze=True).to_dict()

# registryToImmportDict
registryDict = {**vars(STUDY),
                **vars(STUDY_CATEGORIZATION),
                **vars(STUDY_DESIGN),
                **vars(STUDY_DETAILS),
                **vars(STUDY_PUBMED),
                **vars(COD)}

registryDict['SARS-CoV-2_Vaccine_Type'] = vaccine_type

# Looping through ImmPort Template to get the correct order of the 'study' section
for se_number in range(se[0],se[3]):    
 
    if temp_ws["A"][se_number].value != None and registryToImmportDict.get(temp_ws["A"][se_number].value) != None:
#         print(temp_ws["A"][se_number].value)
        
        # Using a mapping key + using info in our classes to map the data
#         print(temp_ws["A"][se_number].value)
        reg_key = registryToImmportDict.get(temp_ws["A"][se_number].value).strip().replace(' ',"_").replace('*',"")
#         print(registryDict.get(reg_key))
        try:
            # If input is a list, we will turn it into a string (since it cant be a list)
            if type(registryDict.get(reg_key)) == list:
                temp_ws["B"][se_number].value = ', '.join(registryDict.get(reg_key))
            else:
                temp_ws["B"][se_number].value = registryDict.get(reg_key)

        except:
            print(f"{reg_key} did not work")
            
#     else:
#         print('blank')


bsd = pd.DataFrame(temp_ws.values).replace({None: '', 'None': ''})
# bsd.to_excel(os.path.join(OUT_DIR, f'PMID{PMID}_study.xlsx'), index=False, header = False)
bsd.to_csv(os.path.join(OUT_DIR,f'{BASIC_STUDY_TEMPLATE}.txt'),
           header = False,
           index = False,
           sep = '\t')

bsd.head(30)

### errors: None

# Protocol 

protocol_ws = load_workbook(PATH_protocols)['protocols.txt']
protocol_ws = seroFxn.remove_excess(protocol_ws)
seroFxn.add_df(protocol_ws, PROTOCOLS, add_header = False, stagger = 1)

protocol_df = pd.DataFrame(protocol_ws.values).replace({None: '', 'None': ''})
protocol_df.to_csv(os.path.join(OUT_DIR,f'{PROTOCOL_TEMPLATE}.txt'),
                   header = False, 
                   index = False,
                   sep = '\t')
try:
    shutil.copy(os.path.join(BASE_DIR,'templated_data', f'{PROTOCOLS.Protocol_Name[1]}.txt'),
                os.path.join(OUT_DIR,f'{PROTOCOLS.Protocol_Name[1]}.txt'))
except:
    print(f'File: {PROTOCOLS.Protocol_Name[1]}.txt does not exist')

#### errors: None


# Experiment

# creating a map of the assay types to the SeroNet descriptors 
reg_description = pd.read_excel(df_path, sheet_name = map_sheet)
descriptions = dict(zip(reg_description['Unnamed: 1'][4:], reg_description['Unnamed: 2'][4:]))

#creating a df to add into the worksheet
Assay_used = STUDY_EXPERIMENTS.Experiment_Assay_Type
experiments_df = pd.DataFrame({
    'Column Name': ['']*len(Assay_used),
    'User Defined ID': [f'PMID{PMID}_exp-0'+str(i+1) for i,k in enumerate(Assay_used)],
    'Name': [Assay_used[i+1] for i, k in enumerate(Assay_used)],
    'Description': [descriptions.get(k) for i, k in enumerate(Assay_used)],
    'Measurement Technique': [STUDY_EXPERIMENTS.Experiment_Assay_Type[i+1] for i, k in enumerate(Assay_used)],
    'Study ID': [STUDY.Study_Identifier]*len(Assay_used),
    'Protocol ID(s)': [PROTOCOLS.Protocol_ID[1]]*len(Assay_used)
})

# loading experiment template and removing excess rows and columns
experiment_ws = load_workbook(PATH_experiments)['experiments.txt']
experiment_ws = seroFxn.remove_excess(experiment_ws)

# adding df to bottom of ws
seroFxn.add_df(experiment_ws, experiments_df, add_header = False)


experiments_df = pd.DataFrame(experiment_ws.values).replace({None: '', 'None': ''})
experiments_df.to_csv(os.path.join(OUT_DIR,f'{EXP_TEMPLATE}.txt'),
                      header = False, 
                      index = False,
                      sep = '\t')

experiments_df

#### errors: None

# REAGENT

NumReagents = len(REAGENTS.Reagent_ID)
empty = [''] * NumReagents

reagent_df = pd.DataFrame({
    'Column Name': empty,
    'User Defined ID': REAGENTS.Reagent_ID,
    'Name': REAGENTS.SARS_CoV_2_Antigen,
    'Description': REAGENTS.Assay_Use,
    'Manufacturer': REAGENTS.Manufacturer,
    'Catalog Number': REAGENTS.Catalog,
    'Lot Number': empty,
    'Weblink': empty,
    'Contact': empty
    
})

# loading experiment template and removing excess rows and columns
reagent_ws = load_workbook(PATH_reagent)['reagents.Other.txt']
reagent_ws = seroFxn.remove_excess(reagent_ws)

# adding df to bottom of ws
seroFxn.add_df(reagent_ws, reagent_df, add_header = False)
reagent_df = pd.DataFrame(reagent_ws.values).replace({None: '', 'None': ''})

# saving df
reagent_df.to_csv(os.path.join(OUT_DIR,f'{REAGENT_TEMPLATE}.txt'),
                   header = False, 
                   index = False,
                   sep = '\t')

reagent_df

### errors: None

# SUBJECTS
## Human

SUBJECT_HUMAN.User_Defined_ID

if SUBJECT_HUMAN:
    print ("SUBJECT_human data")
    species = SUBJECT_HUMAN.User_Defined_ID
    empty = ['']*len(species)
    
    if len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type):
        vaccine_name = [i.split('; ')[0] for i in SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type]
        vaccine_type = [i.split('; ')[1] for i in SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type]
    else:
        vaccine_name = empty
        vaccine_type = empty
        
    
    SUBJECT_human_df = pd.DataFrame({
        'Column Name': empty,
        'Subject ID': [f"PMID{PMID}_human_subject-0{int(i+1)}" for i in range(len(species))],
        'Arm Or Cohort ID': SUBJECT_HUMAN.User_Defined_ID, #I feel like this needs to be defined
        'Gender': SUBJECT_HUMAN.Sex_at_Birth, 
        'Min Subject Age': 0, # add this to validator in dataclass 
        'Max Subject Age': 89, # add this to validator in dataclass 
        'Age Unit': [STUDY_DETAILS.Age_Unit]*len(species),
        'Age Event': SUBJECT_HUMAN.Age_Event, 
        'Age Event Specify': empty,
        'Subject Phenotype': empty, 
        'Subject Location': 'United States of America', # This needs to be changed
        'Ethnicity': SUBJECT_HUMAN.Ethnicity,
        'Race': SUBJECT_HUMAN.Race, 
        'Race Specify': empty,
        'Description': empty, 
        'Result Separator Column': empty, 
        'Exposure Process Reported': 'no exposure', #not sure 
        'Exposure Material Reported': vaccine_name, # change this
        'Exposure Material ID': vaccine_type,
        'Disease Reported': empty,
        'Disease Ontology ID': empty,
        'Disease Stage Reported': empty 
    })
    
    # loading experiment template and removing excess rows and columns
    human_ws = load_workbook(PATH_subject_human)['subjectHumans.txt']
    human_ws = seroFxn.remove_excess(human_ws)

    # adding df to bottom of ws
    seroFxn.add_df(human_ws, SUBJECT_human_df, add_header = False)
    SUBJECT_human_df = pd.DataFrame(human_ws.values).replace({None: '', 'None': ''})

#     # saving df
    SUBJECT_human_df.to_csv(os.path.join(OUT_DIR,f'{SUBJ_HUMAN_TEMPLATE}.txt'),
                       header = False, 
                       index = False,
                       sep = '\t')

#     print(SUBJECT_human_df)
    
else:
    print('no human data')
    


## Organism
# - problems: AOC is something that will be hard to map, depending on what we are doing. We might have to change the sections or add them together?

len(SUBJECT_ORGANISM.COVID_19_Disease_Severity)

len(vaccine_type)

COD.Reported_Health_Condition[0]

COD.Reported_Health_Condition

if SUBJECT_ORGANISM:  # Not sure how this plays out. Might need to do a mock one. Will it be 1 subject per study?
    species = SUBJECT_ORGANISM.User_Defined_ID
    empty = ['']*len(species)
    
    if len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
        vaccine_name = [i.split('; ')[0] for i in SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type]
        vaccine_type = [i.split('; ')[1] for i in SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type]
    else:
        vaccine_name = empty
        vaccine_type = empty
    
    print ("SUBJECT_organism data")
    SUBJECT_organism_df = pd.DataFrame({
        'Column Name':empty,
        'Subject ID': [f"PMID{PMID}_organism_subject-0{int(i+1)}" for i in range(len(species))], 
        'Arm Or Cohort ID': SUBJECT_ORGANISM.User_Defined_ID, #I feel like this needs to be defined
        'Gender': SUBJECT_ORGANISM.Sex_at_Birth,
        'Min Subject Age': [STUDY_DETAILS.Minimum_Age]*len(species),
        'Max Subject Age': [STUDY_DETAILS.Maximum_Age]*len(species),
        'Age Unit': [STUDY_DETAILS.Age_Unit] * len(species),
        'Age Event': SUBJECT_ORGANISM.Age_Event,
        'Age Event Specify': empty,
        'Subject Phenotype': empty, ## This does not exist
        'Subject Location': SUBJECT_ORGANISM.Study_Location, # This needs to be changed. SUBJECT.Study_Location,
        'Species': SUBJECT_ORGANISM.Species,
        'Strain': SUBJECT_ORGANISM.Biosample_Types,
        'Strain Characteristics': SUBJECT_ORGANISM.Strain_Characteristics,
        'Result Separator Column': empty,
        'Exposure Process Reported': SUBJECT_ORGANISM.SARS_CoV2_History,
        'Exposure Material Reported': vaccine_name,
        'Exposure Material ID': vaccine_type, 
        'Disease Reported': [COD.Reported_Health_Condition[0]]*len(species), # also maybe not...
        'Disease Ontology ID': empty,
        'Disease Stage Reported': SUBJECT_ORGANISM.COVID_19_Disease_Severity
    })
    
        # loading experiment template and removing excess rows and columns
    organism_ws = load_workbook(PATH_subject_organism)['subjectAnimals.txt']
    organism_ws = seroFxn.remove_excess(organism_ws)

    # adding df to bottom of ws
    seroFxn.add_df(organism_ws, SUBJECT_organism_df, add_header = False)
    SUBJECT_organism_df = pd.DataFrame(organism_ws.values).replace({None: '', 'None': ''})

    # saving df
    SUBJECT_organism_df.to_csv(os.path.join(OUT_DIR,f'{SUBJ_ORGANISM_TEMPLATE}.txt'),
                       header = False, 
                       index = False,
                       sep = '\t')

    SUBJECT_organism_df
    
#     subjectAnimals
else:
    print("no data")

SUBJECT_organism_df

### Error

# PMIDTEST_subject_organism.txt		Parsing		Template: subjectanimals.txt, Size:591 bytes
# PMIDTEST_subject_organism.txt		Information		Validation Level = Standard
# PMIDTEST_subject_organism.txt	3	Stored in:file_info		Size:591 bytes
# PMIDTEST_subject_organism.txt	4	Information		The combined status for subject_id (PMIDTEST_organism-01) and table subject has values:  subject_required (yes), subject_user_defined_id (PMIDTEST_organism-01), and subject_accession ().
# PMIDTEST_subject_organism.txt	4	Information		The combined status for arm_or_cohort_id () and table arm_or_cohort has values:  arm_or_cohort_required (yes), arm_or_cohort_user_defined_id (), and arm_accession ().
# PMIDTEST_subject_organism.txt	4	Error	Value reported in template does not match preferred vocabulary. See template documentation for preferred vocabulary terms.	The Lookup Table check "lk_species" in field "Species" "Golden Hampster".  The row has ID field(s) "subject_user_defined_id" and value "PMIDTEST_organism-01".
# PMIDTEST_subject_organism.txt	4	Error	Value reported in template does not match preferred vocabulary. See template documentation for preferred vocabulary terms.	The Lookup Table check "lk_subject_location" in field "Subject Location" "Northeast".  The row has ID field(s) "subject_user_defined_id" and value "PMIDTEST_organism-01".
# PMIDTEST_subject_organism.txt	4	Error	The Arm Or Cohort ID must always be predefined	The check uses fields ["arm_or_cohort_required_value", "arm_or_cohort_required"] whose values are [no, yes].  The row has ID field(s) "subject_user_defined_id" and value "PMIDTEST_organism-01".
# PMIDTEST_subject_organism.txt	4	Error	subject age_unit "Years" is not in age_unit list linked to study_accession "[]"	The check-value-in-entity check "Years" in field "study_accession" "".  The row has ID field(s) "subject_user_defined_id" and value "PMIDTEST_organism-01".
# PMIDTEST_subject_organism.txt	4	Error	Missing required value	The Rule check "Required Column" for Field "Arm Or Cohort ID".  The row has ID field(s) "subject_user_defined_id" and value "PMIDTEST_organism-01".
# PMIDTEST_subject_organism.txt	4	Information		Adding User Defined ID "SUB1000000117:ARM1000000118:Infected:N/A::COVID-19:DOID:0080600:Mild" for table "immune_exposure"
# PMIDTEST_subject_organism.txt		Failed		Template: subjectanimals.txt, Size:591 bytes

# ASSESSMENT 
There will be 1 excel document produced per assesment.

# NumAssessments = len(ASSESSMENT.Assessment_ID)
# empty = [''] 


# for i in range(NumAssessments): ### THIS NEEDS TO BE 1 cell to the right. So the actual vaules 
#     assessment_name = ''
#     i += 1
#     if ASSESSMENT.Measured_Behavioral_or_Psychological_Factor[i]:
#         # 'Measured Behavioral or Psychological Factor'
#         assessment_name = ASSESSMENT.Measured_Behavioral_or_Psychological_Factor[i]
#     elif ASSESSMENT.Measured_Social_Factor[i]: 
#         # 'Measured Social Factor'
#         assessment_name = ASSESSMENT.Measured_Social_Factor[i]
#     elif ASSESSMENT.SARS_CoV_2_Symptoms[i]:
#         # 'SARS-CoV-2 Symptoms'
#         assessment_name = ASSESSMENT.SARS_CoV_2_Symptoms[i]
#     else:
#         print('No assessment name')


# # need to figure out named report 
# # not sure where the data goes for the assessments

#     assessment_df = pd.DataFrame({
#         'Column Name': empty,
#         'Subject ID': SUBJECT.Subject_ID,  #assuming 1 subj per study .. treat as a study id
#         'Assessment Panel ID': ASSESSMENT.Assessment_ID[i],
#         'Study ID': STUDY.Study_Identifier,
#         'Name Reported': assessment_name,
#         'Assessment Type': ASSESSMENT.Assessment_Clinical_and_Demographic_Data_Provenance[i],
#         'Status': ASSESSMENT.Assessment_Demographic_Data_Types_Collected[i],
#         'CRF File Names': empty,
#         'Result Separator Column': empty,
#         'User Defined ID': ASSESSMENT.Assessment_ID[i], #
#         'Planned Visit ID': PLANNED_VISIT.User_Defined_ID[i],
#         'Name Reported ': f"component_{assessment_name}", # space might not work
#         'Study Day': i, # Not sure, this says assessment component study day
#         'Age At Onset Reported': empty,
#         'Age At Onset Unit Reported': empty,
#         'Is Clinically Significant': empty,
#         'Location Of Finding Reported': empty,
#         'Organ Or Body System Reported': empty,
#         'Result Value Reported': 'NA',
#         'Result Unit Reported': empty,
#         'Result Value Category': empty,
#         'Subject Position Reported': empty,
#         'Time Of Day': empty,
#         'Verbatim Question': empty,
#         'Who Is Assessed': empty
#     })
    
#     # loading experiment template and removing excess rows and columns
#     assessment_ws = load_workbook(PATH_assessment)['assessments.txt']
#     assessment_ws = seroFxn.remove_excess(assessment_ws)

#     # adding df to bottom of ws
#     seroFxn.add_df(assessment_ws, assessment_df, add_header = False)
#     assessment_df = pd.DataFrame(assessment_ws.values).replace({None: '', 'None': ''})
# #     assessment_df.rename(columns=lambda x: x.strip())

#     # saving file
#     assessment_df.to_csv(os.path.join(OUT_DIR,f'panel0{i}_{ASSESSMENT_TEMPLATE}.txt'),
#                        header = False, 
#                        index = False,
#                        sep = '\t')
# #     print(assessment_df.head())

# class my_class(object):
#     def __init__(self):
#         self.lis1 = []
#         self.dict1 = []

# #     def __nonzero__(self):
# #         return bool(self.lis1 or self.dict1)
    
#     def __len__(self):
#         for field in self.fields():
#             print(field.name, getattr(self, field.name))
            
#         return len(self.lis1) + len(self.dict1)
    
#     def __post_init__(self):
#         for field in self.__dataclass_fields__:
#             value = getattr(self, field)
#             print(field, value)

# obj = my_class()
# # obj.lis1 = [1,2,3,4]
# # obj.dict1 = [1,2,3,4]

# if obj:
#     print ("Available")
# else:
#     print ("Not available")

