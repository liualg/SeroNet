#!/usr/bin/env python
# coding: utf-8

'''
This script is compatibale with Registry Version v1.3+
- This script is the abridged version of the full Registry2ImmPort_full script 
- This will ONLY populate the Basic Study Design, Subject (Human / Organism), Experiment, and Protocol templates

'''

import time
import pandas as pd
import numpy as np
import os, shutil
import pathlib
import inspect
import datetime as dt
from sys import platform
from glob import glob
import sys
import json

from tqdm import tqdm

import argparse 

from dataclasses import dataclass, field
# import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

## Importing Functions and Dataclass
import seronetDataclass as seroClass
import seronetFunctions as seroFxn
import JSONparse_template as pt
import JSONcreateSuggestions as cs

import warnings

warnings.simplefilter("ignore")

if platform == "darwin":
    os.system('clear')
else:
    os.system('cls')


file_type = "json"
DR_NUMBER = "DR51"

#########################################
######### Taking in Inputs ##############
#########################################

def create_short(PMID, user_input_path=False):
    sheet_name = 'SeroNet Registry Template'
    map_sheet = 'Registry Definitions'
    EVS_DICT = 'Seronet_Study_Descriptors_v1.3_EVS.xlsx'
    EVS_sheet = 'EVS Mapping'

    if user_input_path == True:
        # using local file
        path = input("Enter File Path: ")
        df_path = str(path)
        BASE_DIR = pathlib.PurePosixPath(df_path).parent


    else:
        # Using Box directory
        # finding correct Box Base
        if platform == "darwin":
            box_base = "~/Library/CloudStorage/Box-Box/SeroNet Curation/SeroNet Public Data"
        else: 
            print("User has windows")
            box_base = os.path.join("Users",os.getlogin(), "Box")


        #File Paths
        BASE_DIR = seroFxn.get_box_dir(box_base, PMID)
        # print(BASE_DIR)

        try:
            df_path = glob(os.path.join(BASE_DIR,'templated_data',f'PMID{PMID}*eviewed.xlsm'))[0]
        except:
            # print("** NOT using reviewed files **")
            try:
                df_path = glob(os.path.join(BASE_DIR,'templated_data',f'PMID{PMID}*.xlsm'))[0]
            except FileNotFoundError:
                sys.exit("ERROR:: Incorrect Template format. Cannot Find File")
            
        
    # df_path = os.path.join(BASE_DIR,'templated_data', file + ".xlsm")
    print("file path: {}".format(df_path))
    # df_path = os.path.join(BASE_DIR,'templated_data', "PMID34431693_registry.xlsm")

    # dictionary info
    registryToImmportDict_file = os.path.join("dictionary", "registryToBasic.csv")

    registryToImmportDict = pd.read_csv(registryToImmportDict_file,
                                        header=None, 
                                        index_col=0).squeeze().to_dict()

    # ImmPort Templates (link to web?)
    PATH_basic_stdy_template = os.path.join("template", "basic_study_design.xlsx")
    PATH_protocols = os.path.join("template", "protocols.xlsx")
    PATH_experiments = os.path.join("template", "experiments.xlsx") # should be tied to EXP sample + bio sample + subject
    PATH_subject_human = os.path.join("template", "subjectHumans.xlsx")
    PATH_subject_organism = os.path.join("template", "subjectAnimals.xlsx")



    # Automate output... 
    OUT_DIR = os.path.join(BASE_DIR, f'ImmPort_templates_{DR_NUMBER}') 
    # OUT_DIR = './test/'
    PATH_pmid_basic_stdy_template = f'PMID{PMID}_study.xlsx'

    BASIC_STUDY_TEMPLATE = f'PMID{PMID}_basic'
    EXP_TEMPLATE = f'PMID{PMID}_experiments'
    PROTOCOL_TEMPLATE = f'PMID{PMID}_protocol'
    SUBJ_HUMAN_TEMPLATE =  f'PMID{PMID}_subject_human'
    SUBJ_ORGANISM_TEMPLATE =  f'PMID{PMID}_subject_organism'

    # Make Dir if it does not exist
    try:
        os.mkdir(OUT_DIR)
        print(f'\nCreating output directory - {OUT_DIR}')
    except FileExistsError:
        print('\nWill not create directory - already exists')
        pass

    # import file
    book = load_workbook(df_path)
    registry = book[sheet_name]
    registry.delete_cols(1)

    print("\n~~ File Information ~~")

    # In[3]:


    # Class names
    class_names = ['study_pubmed', 'study', 'study_personnel', 'study_file',
                   'study_link', 'study_categorization', 'study_design',
                   'protocol', 'condition_or_disease', 'Intervention Agent',
                   'study_details', 'inclusion_exclusion',
                   'Subject Type: human', 'Subject Type: model organism', 'planned_visit',
                   'Serology Experiments', 'Experiment Samples', 'Experiments'
                  ]



    VARS_TO_CLEAN = ['', 'N/A', 'n/a', 'na', np.nan, None]
    clean_other = VARS_TO_CLEAN + ['Other']

    sp = seroFxn.get_sections(registry, class_names)
    sp.append(200)

    #########################################
    ######        Main Loop       ###########
    #########################################

    # Looping through each section in the Registy template
    for section_number in range(len(sp)-1):
        temp_wb = Workbook()
        temp_ws = temp_wb.active

        #making a temp workbook to store each section. This will be turned into df
        for i in registry.iter_rows(values_only = True,
                                    min_row = sp[section_number]+1,
                                    max_row = sp[section_number+1]-1):
            temp_ws.append(i)
            
        max_row = temp_ws.max_row
        max_col = temp_ws.max_column
        seroFxn.remove_excess(temp_ws)
        
        df = pd.DataFrame(temp_ws.values)
        sub_section = registry.cell(row=sp[section_number], column = 1).value.strip()
        if sub_section == 'study':
            df = seroFxn.edit_df(df)
            
            STUDY = seroClass.study(
                df['Study Identifier'][1],
                df['Study Name'][1],
                df['Publication Title'][1],
                df['Study Objective'][1],
                df['Study Description'][1],
                df['Primary Institution Name'][1]
                )
        
        elif sub_section == 'study_pubmed':
            df = seroFxn.edit_df(df)
            
            STUDY_PUBMED = seroClass.study_pubmed(
                # df['Pubmed ID'][1]
                PMID
            )
            
        elif sub_section == 'study_personnel':
            df = seroFxn.edit_df(df)
            try:
                STUDY_PERSONNEL = seroClass.study_personnel(
                    df['Personnel ID'],
                    df['Honorific'],
                    df['Last Name'],
                    df['First Name'],
                    df['Suffixes'],
                    df['Organization'],
                    df['ORCID ID'],
                    df['Email'],
                    df['Title In Study'],
                    df['Role In Study'],
                    df['Site Name']
                )
            except:
                print("trying older format: Title in Study")
                STUDY_PERSONNEL = seroClass.study_personnel(
                    df['Personnel ID'],
                    df['Honorific'],
                    df['Last Name'],
                    df['First Name'],
                    df['Suffixes'],
                    df['Organization'],
                    df['ORCID ID'],
                    df['Email'],
                    df['SeroNet Title In Study'],
                    df['Role In Study'],
                    df['Site Name']
                )


        elif sub_section == 'study_file':
            df = seroFxn.edit_df(df)

            STUDY_FILE = seroClass.study_file(
                df['Study File Name'],
                df['Study File Description'],
                df['Study File Type']  
            )
            
        
        elif sub_section == 'study_link':
            df = seroFxn.edit_df(df)

            STUDY_LINK = seroClass.study_link(
                df['Link Name'],
                df['Value']
            )
                 
            
        elif sub_section == 'study_categorization':
            df = seroFxn.edit_df(df)

            try:
                STUDY_CATEGORIZATION = seroClass.study_categorization(
                    df['Research Focus*'][1],
                    df['Study Type'][1],
                    df['Keywords'][1],
                )
            except:
                STUDY_CATEGORIZATION = seroClass.study_categorization(
                    df['Research Focus'][1],
                    df['Study Type'][1],
                    df['Keywords'][1],
            )
        
        elif sub_section == 'study_design':
            df = seroFxn.edit_df(df)

            if (df.shape != (2,0)): # checking size. There has to be a better way to do this
                try:
                    STUDY_DESIGN = seroClass.study_design(
                        # df['Clinical Study Design*'],
                        df[df.columns[df.columns.str.startswith('Clinical')]],
                        df[df.columns[df.columns.str.startswith('in silico')]]
                        # df['c Model Type*']
                    )
                except:
                    print("Using older template")
                    STUDY_DESIGN = seroClass.study_design(
                        df['Clinical Study Design'],
                        df['in silico Model Type*']
                    )

        elif sub_section == 'protocol':
            df = seroFxn.edit_df(df)

            PROTOCOLS = seroClass.protocols(
                df['Protocol ID'],
                df['Protocol File Name'],
                df['Protocol Name'],
                # df['Protocol Description'],
                df[df.columns[df.columns.str.startswith('Protocol Descr')][0]],
                df['Protocol Type'],
            )
            
        elif sub_section == 'condition_or_disease':
            df = seroFxn.edit_df(df)

            COD = seroClass.condition_or_disease(
                list(df['Reported Health Condition* '])
            )
            
        elif sub_section == 'Intervention Agent':
            df = seroFxn.edit_df(df)

            INTERVENTION_AGENT = seroClass.Intervention_Agent(
                list(df['SARS-CoV-2 Vaccine Type*'])  ### IF THIS IS NOTHING, change to 'NA'
            )

        elif sub_section == 'study_details':
            df = seroFxn.edit_df(df)
            STUDY_DETAILS = seroClass.study_details(
                df['Clinical Outcome Measure'][1],
                df['Enrollment Start Date'][1],
                df['Enrollment End Date'][1],
                df['Number of Study Subjects'][1],
                df['Age Unit'][1],
                df['Minimum Age'][1],
                df['Maximum Age'][1]
                )

        elif sub_section == 'inclusion_exclusion':
            df = seroFxn.edit_df(df)
            

            INCLUSION_EXCLUSION = seroClass.inclusion_exclusion(
                df['Inclusion ID'],
                df['Inclusion Criterion'],
                df['Inclusion Criterion Category']  
            )
            # print(INCLUSION_EXCLUSION)
            
        elif sub_section == 'Subject Type: human':
            df = seroFxn.edit_df(df)
            tem_was_here = df

            try:
                SUBJECT_HUMAN = seroClass.subject_type_human(
                    df['Arm ID'],
                    df['Arm Name'],
                    df['Study Population Description'],
                    df['Arm Type'],
                    df['Ethnicity*'],
                    df['Race*'],
                    df['Race Specify'],
                    df['Treatment'],
                    df['Sex at Birth*'],
                    df['Age Event'],
                    df['Subject Phenotype'],
                    df['Study Location*'],
                    df['Assessment Name'],
                    df['Measured Behavioral or Psychological Factor*'],
                    df['Measured Social Factor*'],
                    df['SARS-CoV-2 Symptoms*'],
                    df['Assessment_Clinical  and Demographic Data Provenance*'],
                    df['Assessment_Demographic Data Types Collected*'],
                    df['SARS-CoV2 History*'],
                    df['SARS-CoV-2 Vaccine Type*'],
                    df['COVID-19 Disease Severity*'],
                    df['Post COVID-19 Symptoms'],
                    df['COVID-19 Complications']
                    )
            except:
                print("** using older version of subject human **")
                SUBJECT_HUMAN = seroClass.subject_type_human(
                    df['Arm ID'],
                    df['Arm Name'],
                    df['Study Population Description'],
                    df['Arm Type'],
                    df['Ethnicity*'],
                    df['Race*'],
                    df['Race Specify'],
                    df['Description'],
                    df['Sex at Birth*'],
                    df['Age Event'],
                    df['Subject Phenotype'],
                    df['Study Location*'],
                    df['Assessment Name'],
                    df['Measured Behavioral or Psychological Factor*'],
                    df['Measured Social Factor*'],
                    df['SARS-CoV-2 Symptoms*'],
                    df['Assessment_Clinical  and Demographic Data Provenance*'],
                    df[df.columns[df.columns.str.startswith('Assessment_Demographic Data Types')]],
                    df['SARS-CoV2 History*'],
                    df['SARS-CoV-2 Vaccine Type*'],
                    df['COVID-19 Disease Severity*'],
                    df['Post COVID-19 Symptoms'],
                    df['COVID-19 Complications']
                    )

        elif sub_section == 'Subject Type: model organism':
            df = seroFxn.edit_df(df)
            try:
                SUBJECT_ORGANISM = seroClass.subject_type_mode_organism(
                    df['Arm ID'],
                    df['Arm Name'],
                    df['Study Population Description'],
                    df['Arm Type'],
                    df['Genus and Species'],
                    df['Biosample Type'],
                    df['Strain Characteristics'],
                    df['Sex at Birth*'],
                    df['Age Event'],
                    df['Subject Phenotype'],
                    df['Study Location*'],
                    df['SARS-CoV2 History*'],
                    df['SARS-CoV-2 Vaccine Type*'],
                    df['COVID-19 Disease Severity*'],
                    df['Post COVID-19 [PASC] Symptoms'],
                    df['COVID-19 Complications']
                    )
            except:
                print('Using older verion of organism section')
                SUBJECT_ORGANISM = seroClass.subject_type_mode_organism(
                    df['Arm ID'],
                    df['Arm Name'],
                    df['Study Population Description'],
                    df['Arm Type'],
                    df.filter(regex=('.*species.+')),
                    df['Biosample Type'],
                    df['Strain Characteristics'],
                    df['Sex at Birth*'],
                    df['Age Event'],
                    df['Subject Phenotype'],
                    df['Study Location*'],
                    df['SARS-CoV2 History*'],
                    df['SARS-CoV-2 Vaccine Type*'],
                    df['COVID-19 Disease Severity*'],
                    df['Post COVID-19 Symptoms'],
                    df['COVID-19 Complications']
                )
        
        elif sub_section == 'planned_visit':
            df = seroFxn.edit_df(df)

            PLANNED_VISIT = seroClass.planned_visit(
                df['Visit ID'],
                df['Visit Name'],
                df['Visit Order Number'],
                df['Visit Min Start Day'],
                df['Visit Max Start Day'],
                df['Visit Start Rule']                                
        )
        
        elif sub_section == 'Experiment Samples':
            sys.exit("ERROR:: Incorrect Template format. Need to follow v1.2.3")
        #     FileExistsError
        #     df = seroFxn.edit_df(df)

        #     EXPERIMENT_SAMPLES = seroClass.study_experiment_samples(
        #         df['Associated Arm ID'],
        #         df['Associated Planned Visit ID'],
        #         df['Expt Sample User Defined ID'],
        #         df['Biospecimen Type*'],
        #         df['Biospecimen Collection Point*']                             
        # )
        
        elif sub_section in ['Experiments' , 'Serology Experiments']:
            df = seroFxn.edit_df(df)
            try:
                EXPERIMENTS = seroClass.experiments(
                    df['Associated Arm ID(s)'],
                    df['Associated First Planned Visit ID'],
                    df['Assay Type'],
                    df['Experiment Name'],
                    df['Experiment Results File Name'],
                    df['Biospecimen Type*'],
                    df['Biospecimen Collection Point'], 
                    df['SARS-CoV-2 Antigen*'],
                    df['Assay Use'],
                    df['Manufacturer'],
                    df['Catalog #'],
                    df['Virus Target'],
                    df['Antibody Isotype*'],
                    df['Reporting Units'],
                    df['Assay Reporting Format']
            )
            except:
                print("** Using older version of experiments")
                EXPERIMENTS = seroClass.experiments(
                    df['Associated Arm ID(s)'],
                    df['Associated First Planned Visit ID'],
                    df['Assay Type'],
                    df['Experiment Name'],
                    df['Experiment Results File Name'],
                    df['Biospecimen Type*'],
                    df['Biospecimen Collection Point'], 
                    df['SARS-CoV-2 Antigen*'],
                    df['Assay Use'],
                    df['Manufacturer'],
                    df['Catalog Number'],
                    df['Virus Target'],
                    df['Antibody Isotype*'],
                    df['Reporting Units'],
                    df['Assay Reporting Format']
            )

        else:
            print(sub_section, ': does not exist')



    # # Sending back to ImmPort Template #



    ImmPortClassNames = ['study', 'study_categorization', 'study_2_condition_or_disease',
                   'arm_or_cohort','study_personnel', 'planned_visit', 'inclusion_exclusion',
                   'study_2_protocol', 'study_file', 'study_link', 'study_pubmed'] 


    # In[9]:


    shutil.copy(PATH_basic_stdy_template, PATH_pmid_basic_stdy_template)
    basic_stdy_template = load_workbook(PATH_pmid_basic_stdy_template)
    bst_ws = basic_stdy_template['basic_study_design.txt']

    se = seroFxn.get_sections(bst_ws, ImmPortClassNames)

    #########################################
    ######### Basic Study Template ##########
    #########################################
    vaccine_name = []
    # print("*****",SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type)

    if SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type.any() and SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type.any():
        
        vaccine_name, vaccine_type = seroFxn.get_vaccine(set(list(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type) + list(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type)),
                        clean_other)
        
        
    elif SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type.any():
        
        vaccine_name, vaccine_type = seroFxn.get_vaccine(set(list(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type)),
                        clean_other)
                                
    elif SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type.any():
        
        vaccine_name, vaccine_type = seroFxn.get_vaccine(set(list(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type)),
                        clean_other)
        
    else:
        
        vaccine_name = []

    # print(bool(vaccine_name), vaccine_name)
    if len(vaccine_name):
        vaccine_name = list(set([x.strip() for x in vaccine_name]))

    if SUBJECT_HUMAN and SUBJECT_ORGANISM:
        print('Both human and model organism are used')
        usr_id = list(SUBJECT_HUMAN.User_Defined_ID) + list(SUBJECT_ORGANISM.User_Defined_ID)
        name = list(SUBJECT_HUMAN.Name) + list(SUBJECT_ORGANISM.Name)
        description = list(SUBJECT_HUMAN.Description) + list(SUBJECT_ORGANISM.Description)
        type_report = list(SUBJECT_HUMAN.Type_Reported) + list(SUBJECT_ORGANISM.Type_Reported)
        
        if set(SUBJECT_HUMAN.User_Defined_ID).intersection(SUBJECT_ORGANISM.User_Defined_ID):
            print("\n*** ERROR: Cannot have same User Defined ID's for AOCs:")
            print(set(SUBJECT_HUMAN.User_Defined_ID).intersection(SUBJECT_ORGANISM.User_Defined_ID))
            
    if SUBJECT_HUMAN and not SUBJECT_ORGANISM:
        print('Only human is used')
        usr_id = list(SUBJECT_HUMAN.User_Defined_ID)
        name = list(SUBJECT_HUMAN.Name)
        description = list(SUBJECT_HUMAN.Description)
        type_report = list(SUBJECT_HUMAN.Type_Reported)
        
    if SUBJECT_ORGANISM and not SUBJECT_HUMAN:
        print('Only organism is used')
        usr_id = list(SUBJECT_ORGANISM.User_Defined_ID)
        name = list(SUBJECT_ORGANISM.Name)
        description = list(SUBJECT_ORGANISM.Description)
        type_report = list(SUBJECT_ORGANISM.Type_Reported)



    temp_wb = Workbook()
    temp_ws = temp_wb.active

    #making a temp workbook to store each section. This will be turned into df

    # This is the first part of the df
    for i in bst_ws.iter_rows(values_only = True,
                              max_row = se[2]+1):
        temp_ws.append(i)
        
    # Starting at Arm or Cohort  
    # print(usr_id, name, description, type_report)
    AOC = seroClass.arm_or_cohort(
        list(filter(None, usr_id)),
        list(filter(None, name)),
        list(filter(None, description)),
        list(filter(None, type_report))
    )

    # temp_ws.append(['User Defined ID', 'Name','Description','Type Reported'])
    seroFxn.add_df(temp_ws, AOC)

    seroFxn.add_df(temp_ws, STUDY_PERSONNEL)

    # Checking to see if assessment is used. If it is, then we will add another planned visit ID to 
    # The planned visit section 

    # This area of code need to be updated 
    if SUBJECT_HUMAN and (SUBJECT_HUMAN.Assessment_Name.any() or \
    SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type.any() or \
    SUBJECT_HUMAN.SARS_CoV2_History.any() or \
    SUBJECT_HUMAN.SARS_CoV_2_Symptoms.any() or \
    SUBJECT_HUMAN.Measured_Social_Factor.any() or \
    SUBJECT_HUMAN.Measured_Behavioral_or_Psychological_Factor.any() or \
    SUBJECT_HUMAN.Assessment_Demographic_Data_Types_Collected.any() or \
    SUBJECT_HUMAN.Assessment_Clinical_and_Demographic_Data_Provenance.any()) and \
    'Not Applicable' not in SUBJECT_HUMAN.Assessment_Name:

        add_index = len(PLANNED_VISIT.Name) + 1

        if PLANNED_VISIT.Order_Number.any():
            order_index = PLANNED_VISIT.Order_Number[add_index-1] + 1
        else:
            order_index = 1

        PLANNED_VISIT.User_Defined_ID[add_index] = f'PMID{PMID}_assessment_recorded_pv'
        PLANNED_VISIT.Name[add_index] = 'Visit where an assessment is recorded'
        PLANNED_VISIT.Order_Number[add_index] = order_index
        PLANNED_VISIT.Min_Start_Day[add_index] = 0
        PLANNED_VISIT.Max_Start_Day[add_index] = ''
        PLANNED_VISIT.Start_Rule[add_index] = ''

    seroFxn.add_df(temp_ws, PLANNED_VISIT)
    seroFxn.add_df(temp_ws, INCLUSION_EXCLUSION)

    # # Add protocol
    temp_ws.append([])
    temp_ws.append([PROTOCOLS.ImmPortNAME])
    temp_ws.append(['Protocol ID', PROTOCOLS.Protocol_Name[1]])

    seroFxn.add_df(temp_ws, STUDY_FILE)
    seroFxn.add_df(temp_ws, STUDY_LINK)

    seroFxn.add_df(temp_ws, STUDY_PUBMED)


    # #### Filling in Study Info Sections 



    registryToImmportDict = pd.read_csv(registryToImmportDict_file, 
                                        header=None, 
                                        index_col=0).squeeze().to_dict()
    # registryToImmportDict
    registryDict = {**vars(STUDY),
                    **vars(STUDY_CATEGORIZATION),
                    **vars(STUDY_DESIGN),
                    **vars(STUDY_DETAILS),
                    **vars(STUDY_PUBMED),
                    **vars(COD)}

    registryDict['SARS-CoV-2_Vaccine_Type'] = ['N/A']

    if len([x for x in vaccine_name if x not in VARS_TO_CLEAN]) != 0:
            registryDict['SARS-CoV-2_Vaccine_Type'] = [x for x in vaccine_name if x not in VARS_TO_CLEAN]


            registryDict['SARS-CoV-2_Vaccine_Type'] = list(set((' | '.join(registryDict['SARS-CoV-2_Vaccine_Type'])).split(' | ')))


            # print('########',"\n",registryDict['SARS-CoV-2_Vaccine_Type'])

    # Looping through ImmPort Template to get the correct order of the 'study' section
    for se_number in range(se[0],se[3]):    
     
        if temp_ws["A"][se_number].value != None and registryToImmportDict.get(temp_ws["A"][se_number].value) != None:
            
            # Using a mapping key + using info in our classes to map the data
            reg_key = registryToImmportDict.get(temp_ws["A"][se_number].value).strip().replace(' ',"_").replace('*',"")
            try:
                # If input is a list, we will turn it into a string (since it cant be a list)
                if type(registryDict.get(reg_key)) == list:
                    temp_ws["B"][se_number].value = ', '.join(registryDict.get(reg_key))
                else:
                    temp_ws["B"][se_number].value = registryDict.get(reg_key)

            except:
                print(f"{reg_key} did not work")
                


    bsd = pd.DataFrame(temp_ws.values).replace({None: '', 'None': ''})
    # bsd.to_excel(os.path.join(OUT_DIR, f'PMID{PMID}_study.xlsx'), index=False, header = False)
    bsd.to_csv(os.path.join(OUT_DIR,f'{BASIC_STUDY_TEMPLATE}.txt'),
               header = False,
               index = False,
               sep = '\t')


    #########################################
    ############## Protocol #################
    #########################################

    #Load Protocol
    protocol_ws = load_workbook(PATH_protocols)['protocols.txt']
    protocol_ws = seroFxn.remove_excess(protocol_ws)
    seroFxn.add_df(protocol_ws, PROTOCOLS, add_header = False, stagger = 1)



    protocol_df = pd.DataFrame(protocol_ws.values).replace({None: '', 'None': ''})
    protocol_df.to_csv(os.path.join(OUT_DIR,f'{PROTOCOL_TEMPLATE}.txt'),
                       header = False, 
                       index = False,
                       sep = '\t')
    try:
        shutil.copy(os.path.join(BASE_DIR,'templated_data', f'{PROTOCOLS.Protocol_Name[1]}.txt'),
                    os.path.join(OUT_DIR,f'{PROTOCOLS.Protocol_Name[1]}.txt'))
    except:
        print(f'** File:: {PROTOCOLS.Protocol_Name[1]}.txt does not exist **')


    #########################################
    ##########  EXPERIMENT SAMPLE   #########
    #########################################
    '''
    # biosample should depend on the Planned Visit 
    # if the planned visit is the same, then the biosample id is the same. 

    # Experiment ID should depend in the Assay Type. All Assay types should be the same experimmnet ID

    # Experiment Sample ID should be unique for each case? 

    # Reagent ID should either be no_reagents (if not serology) or the Serology reagent ID 

    # Treatment ID should be "no_sars-cov-2_treatments"

    # Result File Name: resultsNotCurated.txt => ExperimentalDataInStudyFilesTab.txt

    # StudyTimeCollected should link back to the planned_visit.User_Defined_ID + planned_visit.Min_Start_Day
    '''
    reagentID = []
    if (EXPERIMENTS):
        print('yes')
    else:
        print('no')

    if  EXPERIMENTS:
        # creating a map of the assay types to the SeroNet descriptors
        reg_description = pd.read_excel(os.path.join('.','dictionary',EVS_DICT), sheet_name = EVS_sheet)
        descriptions = dict(zip(reg_description['NCIt PT'], reg_description['NCIt Def']))
        
        #dictionary created for visit ID and min day
        studyTime = dict(zip(PLANNED_VISIT.User_Defined_ID, PLANNED_VISIT.Min_Start_Day))

        fillLen =len(EXPERIMENTS.Assay_Type)
        empty = ['']*fillLen
        # print(len(experimentName), len(experimentReportingFormat),len(empty))
        experiment_df = pd.DataFrame({
            'Column Name':empty,
            'User Defined ID': [f'PMID{PMID}_experimentID-0{n+1}' for n in range(fillLen)],
            'Name': EXPERIMENTS.Experiment_Name,
            'Description':empty, #[descriptions.get(assay) for assay in EXPERIMENTS.Assay_Type],
            'Measurement Technique':EXPERIMENTS.Assay_Type,
            'Study ID':[STUDY.Study_Identifier]*fillLen,
            'Protocol ID(s)': [PROTOCOLS.Protocol_ID[1]]*fillLen # This needs to be populated first and then we can populate 'Additional Result File Names'
        })

        exp_ws = load_workbook(PATH_experiments)['experiments.txt']
        exp_ws = seroFxn.remove_excess(exp_ws)

        # adding df to bottom of ws
        seroFxn.add_df(exp_ws, experiment_df, add_header = False)
        experimentSamples_df = pd.DataFrame(exp_ws.values).replace({None: '', 'None': ''})

        # saving df
        experimentSamples_df.to_csv(os.path.join(OUT_DIR,f'{EXP_TEMPLATE}.txt'),
                           header = False, 
                           index = False,
                           sep = '\t')

    """
    #########################################
    ###         SUBJECT: HUMAN       ########
    #########################################
    # - if human, use human
    # - if animal, use animal sheet 
    # - use length as a predictor

    HUMAN SHOULD BE FOR HUMAN CELLS LINES. DO NOT PUT IN ORGANISM

    """


    if SUBJECT_HUMAN:
        # race_specificty =[]
        # for i in SUBJECT_HUMAN.Race: RACE OTHER
        #     if i == 'Other':
        #         race_specificty.append("Other")
        #     else:
        #         race_specificty.append("")


        species = SUBJECT_HUMAN.User_Defined_ID
        empty = ['']*len(species)

        if SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type.any():
            vaccine_name, vaccine_type = seroFxn.get_vaccine(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type, VARS_TO_CLEAN)

        else:
            vaccine_name = empty
            vaccine_type = empty

        SUBJECT_human_df = pd.DataFrame({
            'Column Name': empty,
            'Subject ID': [f"PMID{PMID}_human_subject-0{int(i+1)}" for i in range(len(species))],
            'Arm Or Cohort ID': SUBJECT_HUMAN.User_Defined_ID, #I feel like this needs to be defined
            'Gender': SUBJECT_HUMAN.Sex_at_Birth, 
            'Min Subject Age': [STUDY_DETAILS.Minimum_Age]*len(species), # #ASSUMPTION - defualts to 0
            'Max Subject Age': [STUDY_DETAILS.Maximum_Age]*len(species), # #ASSUMPTION - defaults to 89
            'Age Unit': [STUDY_DETAILS.Age_Unit]*len(species),
            'Age Event': SUBJECT_HUMAN.Age_Event, 
            'Age Event Specify': empty,
            'Subject Phenotype': empty, 
            'Subject Location': SUBJECT_HUMAN.Study_Location,
            'Ethnicity': SUBJECT_HUMAN.Ethnicity,
            'Race': SUBJECT_HUMAN.Race, 
            'Race Specify': SUBJECT_HUMAN.Race_Specify,
            'Description': empty, 
            'Result Separator Column': empty, 
            'Exposure Process Reported': ['unknown']*len(species) , #not sure 
            'Exposure Material Reported': vaccine_name, # change this
            'Exposure Material ID': vaccine_type,
            'Disease Reported': [COD.Reported_Health_Condition[0]]*len(species), # This was not correct
            'Disease Ontology ID': empty,
            'Disease Stage Reported': SUBJECT_HUMAN.COVID_19_Disease_Severity #CHANGE
        })
        
        # loading experiment template and removing excess rows and columns
        human_ws = load_workbook(PATH_subject_human)['subjectHumans.txt']
        human_ws = seroFxn.remove_excess(human_ws)

        # adding df to bottom of ws
        seroFxn.add_df(human_ws, SUBJECT_human_df, add_header = False)
        SUBJECT_human_df = pd.DataFrame(human_ws.values).replace({None: '', 'None': ''})

    #     # saving df
        # SUBJECT_human_df.to_csv(os.path.join(OUT_DIR,f'{SUBJ_HUMAN_TEMPLATE}.txt'),
        #                    header = False, 
        #                    index = False,
        #                    sep = '\t')

    #     print(SUBJECT_human_df)
        print ("SUBJECT: human data created")
    # else:
    #     print('no human data')
        



    """
    #########################################
    #####       SUBJECT: ORGANISM    ########
    #########################################
    For Syrain Hamster: https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?name=Syrian+hamsters 

    Use latin name -- or look at lk species
    https://www.immport.org/shared/templateDocumentation?tab=2&table=lk_species


    """


    if SUBJECT_ORGANISM:  # Not sure how this plays out. Might need to do a mock one. Will it be 1 subject per study?
        species = SUBJECT_ORGANISM.User_Defined_ID
        empty = ['']*len(species)

        if SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type.any():
            vaccine_name, vaccine_type = seroFxn.get_vaccine(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type, VARS_TO_CLEAN)

        else:
            vaccine_name = empty
            vaccine_type = empty

        # print ("SUBJECT_organism data")
        SUBJECT_organism_df = pd.DataFrame({
            'Column Name':empty,
            'Subject ID': [f"PMID{PMID}_organism_subject-0{int(i+1)}" for i in range(len(species))], #SUBJECT_ORGANISM.User_Defined_ID, #
            'Arm Or Cohort ID': SUBJECT_ORGANISM.User_Defined_ID, #I feel like this needs to be defined
            'Gender': SUBJECT_ORGANISM.Sex_at_Birth,
            'Min Subject Age': [STUDY_DETAILS.Minimum_Age]*len(species),
            'Max Subject Age': [STUDY_DETAILS.Maximum_Age]*len(species),
            'Age Unit': [STUDY_DETAILS.Age_Unit] * len(species),
            'Age Event': SUBJECT_ORGANISM.Age_Event,
            'Age Event Specify': empty,
            'Subject Phenotype': empty, ## This does not exist
            'Subject Location': SUBJECT_ORGANISM.Study_Location,
            'Species': SUBJECT_ORGANISM.Species,
            'Strain': SUBJECT_ORGANISM.Biosample_Types,
            'Strain Characteristics': SUBJECT_ORGANISM.Strain_Characteristics,
            'Result Separator Column': empty,
            'Exposure Process Reported': ['unknown']*len(species),
            'Exposure Material Reported': vaccine_name,
            'Exposure Material ID': vaccine_type, 
            'Disease Reported': [COD.Reported_Health_Condition[0]]*len(species), # also maybe not...
            'Disease Ontology ID': empty,
            'Disease Stage Reported': SUBJECT_ORGANISM.COVID_19_Disease_Severity
        })
        
            # loading experiment template and removing excess rows and columns
        organism_ws = load_workbook(PATH_subject_organism)['subjectAnimals.txt']
        organism_ws = seroFxn.remove_excess(organism_ws)

        # adding df to bottom of ws
        seroFxn.add_df(organism_ws, SUBJECT_organism_df, add_header = False)
        SUBJECT_organism_df = pd.DataFrame(organism_ws.values).replace({None: '', 'None': ''})

        # saving df
        # SUBJECT_organism_df.to_csv(os.path.join(OUT_DIR,f'{SUBJ_ORGANISM_TEMPLATE}.txt'),
        #                    header = False, 
        #                    index = False,
        #                    sep = '\t')

        # SUBJECT_organism_df



    #######################################################
    #####      POST: copy log + JSON + move files     #####
    #######################################################
    

    CD = os.getcwd()
    today = dt.datetime.today().strftime('%Y_%m_%d')

    try:
        os.mkdir(os.path.join(CD,BASE_DIR,"log"))
        print(f'Creating log')
    except FileExistsError:
        print('Will not create log - already exists')
        pass

    shutil.copyfile(os.path.join(CD,"log",f"Registry_{today}.log"), os.path.join(CD,BASE_DIR,"log",f"Registry_{today}.log"))
    # shutil.copyfile(os.path.join(CD,"template","xImmPortFillerDocuments","ExperimentalDataInStudyFilesTab.txt"), os.path.join(OUT_DIR,"ExperimentalDataInStudyFilesTab.txt"))
    shutil.copy(df_path, OUT_DIR)

    for filename in glob(os.path.join(CD,BASE_DIR,"submitted_data", '*.*')):
        shutil.copy(filename, OUT_DIR)

    print("\n")
    for i in glob(os.path.join(OUT_DIR,"*.txt")):
        print(os.path.basename(i))
    os.remove(PATH_pmid_basic_stdy_template)

    # while input('Can I clear the screen: (y/n)') != 'y':
    #     time.sleep(1)
    # os.system('clear')
    print("\n\n\n")

    # CREATING JSON
    output_file = os.path.join(OUT_DIR, f'PMID{PMID}_JSON.{file_type}')
    df = pd.read_excel(df_path, sheet_name = 0, header=None)
    df.index += 1
    template = {}
    pt.parse_registry_template(df, template)

    f = open(output_file, "w")
    print(json.dumps(template, indent=4), file = f)
    f.close()

    # ## CREATING SUGGESTIONS
    # cs.add_NLKsuggestions(OUT_DIR, file_type)

    # shutil.copyfile(output_file, os.path.join('.','DR46','JSON',f'PMID{PMID}_JSON.{file_type}'))

