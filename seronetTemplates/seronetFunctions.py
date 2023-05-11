#!/usr/bin/env python
# coding: utf-8

'''
This script is compatibale with Registry Version v.1.1.0
This script is compatibale with Registry Version v.1.2.2
    - Please look at other template 
    - Added '*' to SARS-CoV-2 Antigen* (row 163, column B)
This script is compatibale with Registry Version v.1.2.3

'''

import pandas as pd 
import numpy as np
import os, shutil
import inspect

from tqdm import tqdm

from dataclasses import dataclass, field
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# find file dir to save files base off PMID
def check_input(pmid):
    if len(pmid) == 8:
        return pmid
    
    else:
        try:
            return check_PMID(pmid.split("_")[1])
        except:
            print('PMID format invalid')
            return False

    
def get_box_dir(box_dir, pmid):
    box_base = os.path.abspath(os.path.expanduser(os.path.expandvars(box_dir)))
    depth = 3

    for dirpath, dirnames, filenames in os.walk(box_base):
        if dirpath[len(box_base):].count(os.sep) < depth:
            if "PMID_"+pmid in dirnames:
                return(os.path.join(dirpath,"PMID_"+pmid))


# Return start index for each section in the SeroNet Registry Template 
def get_sections(ws, class_names):
    temp = []
    for i in range(1,ws.max_row):
        if ws["A"][i].value in class_names:
            temp.append(i+1)
    return temp

def column_header_index(sheet):
    count = 0
    row_index = 1

    for row in sheet.iter_rows():
        if [cell.value for cell in row] == [None]*sheet.max_column:
            break
        row_index += 1
    
    return row_index

def row_header_index(sheet):
    count = 0
    col_index = 1

    for col in sheet.iter_cols():
        if [cell.value for cell in col] == [None]*sheet.max_row:
            break
        col_index += 1
    
    return col_index

def remove_excess(sheet):
    sheet.delete_rows(column_header_index(sheet),sheet.max_row)
    sheet.delete_cols(row_header_index(sheet),sheet.max_column)
    return sheet

def edit_df(df_temp):
    df_temp = df_temp.transpose()
    df_temp.columns = df_temp.iloc[0]
    return(df_temp.drop(df_temp.index[0]))

def clean_array(the_array, VARS_TO_CLEAN):
    return [x for x in the_array if x not in VARS_TO_CLEAN]

def remove_whitespace(obj):
    return obj.replace('\n', ' ').replace('\t', ' ').replace('  ', ' ')

def cleanData(s):
    """Removes characters in the input string that will corrupt the final JSON object"""
    if pd.isna(s):
        return ""
    else:
        if isinstance(s, str):
            r1 = re.compile("\n|\t|\r")
            r2 = re.compile('"')
            s = r1.sub(" ", s)
            s = r2.sub('\\"', s)
            s = s.strip()
        elif isinstance(s, datetime):
            s = s.strftime('%Y-%m-%d')
        else:
            pass
        return s

def get_vaccine(arrays, VARS_TO_CLEAN):
    empty = ['']
    vaccine_type = []
    vaccine_name = []
    if len(arrays): #splitting correctly
        # try: #check if there are multiple vaccine names in this section
        # print("&&&&",type(arrays.astype("string")), arrays.astype("string"))
        # print(" ".join(list(arrays)))
        if "|" in " ".join(arrays):
        
            for vaccines_per_sub in arrays:
                v_type = []
                v_name = []
                
                # print(vaccines_per_sub)
                if vaccines_per_sub.lower() not in VARS_TO_CLEAN+['other']: #checking to see if there is NA
                    # print(vaccines_per_sub)
                    for each in vaccines_per_sub.split('|'):
                        if each.strip().lower() not in VARS_TO_CLEAN+['other']:
                            v_type.append(each.split(";")[1].strip())
                            v_name.append(each.split(";")[0].strip())
                        else:
                            v_type.append('')
                            v_name.append(each)

                    vaccine_type.append(' | '.join(v_type))
                    vaccine_name.append(' | '.join(v_name))
                
                else:
                    # print('#########')
                    # print(vaccines_per_sub)
                    vaccine_type.append(vaccines_per_sub)
                    vaccine_name.append(vaccines_per_sub)
        # except:
        else:
            # print('############')
            # print(arrays)
            for i in arrays:
                # print('############')
                # print(f"*{i}*", type(i))
                if i.lower().strip() == "other" or i.lower().strip() == "n/a":
                    vaccine_name.append(i.lower().strip())
                    vaccine_type.append(i.lower().strip())
                else:
                    # print(i.split(';'))
                    vaccine_name.append(i.split(';')[0].strip())
                    vaccine_type.append(i.split(';')[1].strip())
    else:
        vaccine_name = []
        vaccine_type = []

    return vaccine_name, vaccine_type

def capitalize_proper(original_str, filler_words):

    result = ""
    # Split the string and get all words in a list
    list_of_words = original_str.split()
    # Iterate over all elements in list
    for elem in list_of_words:
        # capitalize first letter of each word and add to a string
        if len(result) > 0:
            if elem.strip().lower() not in filler_words:
                result = result + " " + elem.strip().capitalize()
            else:
                result = result + " " + elem.lower().strip()
        else:
            if elem.strip().lower() not in filler_words:
                result = elem.capitalize()
            else:
                result = result + " " + elem.lower().strip()
    # If result is still empty then return original string else returned capitalized.
    if not result:
        return original_str
    else:
        return result


def replace_delimiter(original_str, new_delimiter = " | "):
    # print(original_str)
    if isinstance(original_str, list):
        for i, k in enumerate(original_str):
            # print("########## ASdAS")
            # print(i,k)
            original_str[i] = remove_whitespace(k).replace(" I ", new_delimiter).replace("; ", new_delimiter).replace(" i ", new_delimiter).replace(",", new_delimiter).replace("  ", " ") #replace(" ; ", new_delimiter).replace(".", "")
    else: # .replace(" ; ", new_delimiter)
        original_str = remove_whitespace(original_str).replace(" I ", new_delimiter).replace("; ", new_delimiter).replace(" i ", new_delimiter).replace(",", new_delimiter).replace("  ", " ")

    return original_str 
#################  Adding sections back to dataframes  ##################

def add_df(ws, input_df_class, add_header = True, stagger = 0):
    
    if isinstance(input_df_class, pd.DataFrame): #checking to see if it is a class
        df = input_df_class

    else: #if its a dataframe - just overwrite df to be the dataframe
                
        if input_df_class.ImmPortNAME == 'study_pubmed':
            df = pd.DataFrame({**vars(input_df_class)}, index=[0]).drop(['ImmPortNAME'], axis = 1)
        else:
            df = pd.DataFrame({**vars(input_df_class)}).drop(['ImmPortNAME'], axis = 1)
        
    df.columns = [i.replace("_"," ") for i in df.columns]

    if add_header == True:
        ws.append([])
        ws.append([input_df_class.ImmPortNAME])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(['']*stagger + r)
    else:
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(['']*stagger + r)
        









    