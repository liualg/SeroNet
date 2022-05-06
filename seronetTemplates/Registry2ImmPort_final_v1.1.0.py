#!/usr/bin/env python
# coding: utf-8

# This script is compatibale with Registy Version v.1.1.0

import pandas as pd
import numpy as np
import os, shutil
import inspect
import datetime as dt
from sys import platform

from tqdm import tqdm

import tkinter as tk
from tkinter import simpledialog

from dataclasses import dataclass, field
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

## Importing Functions and Dataclass
import seronetDataclass_v1_1_0 as seroClass
import seronetFunctions as seroFxn

import warnings

warnings.simplefilter("ignore")

try:
    os.system('clear')
except:
    os.system('cls')


# In[2]:


ROOT = tk.Tk()


ROOT.tk.eval(f'tk::PlaceWindow {ROOT._w} center')
ROOT.withdraw()

# the input dialog
USER_INP = simpledialog.askstring(title="PMID",
                                  prompt="Enter PMID:",
                                  parent = ROOT)

#TEST PMIDS
# PMID = "33184236"
# PMID = "34270939"
# PMID = "34431693"
# " 33035201"

PMID = seroFxn.check_input(USER_INP)

# Registry
sheet_name = 'SeroNet Registry Template'
map_sheet = 'Registry Definitions'


# finding correct Box Base
if platform == "darwin":
    box_base = "~/Library/CloudStorage/Box-Box/SeroNet Public Data"
else: 
    print("User has windows")
    box_base = os.path.join( "Users",os.getlogin(), "Box")


#File Paths
BASE_DIR = seroFxn.get_box_dir(box_base, PMID)


try: 
    file = f"PMID{PMID}_v1.1.0"
except:
    file = f"PMID{PMID}_Registry"
    
    
df_path = os.path.join(BASE_DIR,'templated_data', file + ".xlsm")
print("file path: {}".format(df_path))
# df_path = os.path.join(BASE_DIR,'templated_data', "PMID34431693_registry.xlsm")

# dictionary info
registryToImmportDict_file = os.path.join("dictionary", "registryToBasic.csv")

registryToImmportDict = pd.read_csv(registryToImmportDict_file,
                                    header=None, 
                                    index_col=0, 
                                    squeeze=True).to_dict()

# ImmPort Templates (link to web?)
PATH_basic_stdy_template = os.path.join("template", "basic_study_design.xlsx")
PATH_protocols = os.path.join("template", "protocols.xlsx")
PATH_experiments = os.path.join("template", "experiments.xlsx")
PATH_reagent = os.path.join("template", "reagents.Other.xlsx")
PATH_assessment = os.path.join("template", "assessments.xlsx")
PATH_subject_human = os.path.join("template", "subjectHuman.xlsx")
PATH_subject_organism = os.path.join("template", "subjectAnimals.xlsx")


# Automate output... 
OUT_DIR = os.path.join(BASE_DIR, 'ImmPort_templates') 
# OUT_DIR = './33184236_test/'
PATH_pmid_basic_stdy_template = f'PMID{PMID}_study.xlsx'

BASIC_STUDY_TEMPLATE = f'PMID{PMID}_basic'
EXP_TEMPLATE = f'PMID{PMID}_experiments'
PROTOCOL_TEMPLATE = f'PMID{PMID}_protocol'
REAGENT_TEMPLATE = f'PMID{PMID}_reagent'
ASSESSMENT_TEMPLATE =  f'PMID{PMID}_assessment'
SUBJ_HUMAN_TEMPLATE =  f'PMID{PMID}_subject_human'
SUBJ_ORGANISM_TEMPLATE =  f'PMID{PMID}_subject_organism'

# Make Dir if it does not exist
try:
    os.mkdir(OUT_DIR)
    print(f'Creating output directory - {OUT_DIR}')
except FileExistsError:
    print('Will not create directory - already exists')
    pass

# import file
book = load_workbook(df_path)
registry = book[sheet_name]
registry.delete_cols(1)

print("\n~~ File Information ~~")

# In[3]:


# Class names
class_names = ['study_pubmed', 'study', 'study_personnel', 'study_file',
               'study_link', 'study_categorization', 'study_design',
               'protocol', 'condition_or_disease', 'Intervention Agent',
               'study_details', 'inclusion_exclusion',
               'Subject Type: human', 'Subject Type: model organism', 'planned_visit',
               'Study Experiments', 'Study Experiment Samples', 'Reagent', 
               'Results for Serology Assays'
              ]


VARS_TO_CLEAN = ['', 'N/A', 'n/a', np.nan, None]


# In[4]:


sp = seroFxn.get_sections(registry, class_names)


# ### Main Loop

# In[5]:


# Looping through each section in the Registy template
for section_number in range(len(sp)-1):
    temp_wb = Workbook()
    temp_ws = temp_wb.active

    #making a temp workbook to store each section. This will be turned into df
    for i in registry.iter_rows(values_only = True,
                                min_row = sp[section_number]+1,
                                max_row = sp[section_number+1]-1):
        temp_ws.append(i)
        
    max_row = temp_ws.max_row
    max_col = temp_ws.max_column
    seroFxn.remove_excess(temp_ws)
    
    df = pd.DataFrame(temp_ws.values)
    sub_section = registry.cell(row=sp[section_number], column = 1).value.strip()
    
    if sub_section == 'study':
        df = seroFxn.edit_df(df)
        
        STUDY = seroClass.study(
            df['Study Identifier'][1],
            df['Study Name'][1],
            df['Publication Title'][1],
            df['Study Objective'][1],
            df['Study Description'][1],
            df['Primary Institution Name'][1]
            )
    
    elif sub_section == 'study_pubmed':
        df = seroFxn.edit_df(df)
        
        STUDY_PUBMED = seroClass.study_pubmed(
            df['Pubmed ID'][1]
        )
        
    elif sub_section == 'study_personnel':
        df = seroFxn.edit_df(df)

        STUDY_PERSONNEL = seroClass.study_personnel(
            df['Personnel ID'],
            df['Honorific'],
            df['Last Name'],
            df['First Name'],
            df['Suffixes'],
            df['Organization'],
            df['ORCID ID'],
            df['Email'],
            df['Title In Study'],
            df['Role In Study'],
            df['Site Name']
        )


    elif sub_section == 'study_file':
        df = seroFxn.edit_df(df)

        STUDY_FILE = seroClass.study_file(
            df['Study File Name'],
            df['Study File Description'],
            df['Study File Type']  
        )
        
    
    elif sub_section == 'study_link':
        df = seroFxn.edit_df(df)

        STUDY_LINK = seroClass.study_link(
            df['Link Name'],
            df['Value']
        )
             
        
    elif sub_section == 'study_categorization':
        df = seroFxn.edit_df(df)

        STUDY_CATEGORIZATION = seroClass.study_categorization(
            df['Research Focus*'][1],
            df['Study Type'][1],
            df['Keywords'][1],
        )
    
    elif sub_section == 'study_design':
        df = seroFxn.edit_df(df)

        if (df.shape != (2,0)): # checking size. There has to be a better way to do this

            STUDY_DESIGN = seroClass.study_design(
                df['Clinical Study Design'],
                df['in silico Model Type*']
            )
        else:
            study_des = study_design()

    elif sub_section == 'protocol':
        df = seroFxn.edit_df(df)

        PROTOCOLS = seroClass.protocols(
            df['Protocol ID'],
            df['Protocol File Name'],
            df['Protocol Name'],
            df['Protocol Description'],
            df['Protocol Type'],
        )
        
    elif sub_section == 'condition_or_disease':
        df = seroFxn.edit_df(df)

        COD = seroClass.condition_or_disease(
            list(df['Reported Health Condition* '])
        )
        
    elif sub_section == 'Intervention Agent':
        df = seroFxn.edit_df(df)

        INTERVENTION_AGENT = seroClass.Intervention_Agent(
            list(df['SARS-CoV-2 Vaccine Type*'])  ### IF THIS IS NOTHING, change to 'NA'
        )

    elif sub_section == 'study_details':
        df = seroFxn.edit_df(df)
        STUDY_DETAILS = seroClass.study_details(
            df['Clinical Outcome Measure'][1],
            df['Enrollment Start Date'][1],
            df['Enrollment End Date'][1],
            df['Number of Study Subjects'][1],
            df['Age Unit'][1],
            df['Minimum Age'][1],
            df['Maximum Age'][1]
            )

    elif sub_section == 'inclusion_exclusion':
        df = seroFxn.edit_df(df)

        INCLUSION_EXCLUSION = seroClass.inclusion_exclusion(
            df['Inclusion ID'],
            df['Inclusion Criterion'],
            df['Inclusion Criterion Category']  
        )
        
    elif sub_section == 'Subject Type: human':
        df = seroFxn.edit_df(df)
        tem_was_here = df

        SUBJECT_HUMAN = seroClass.subject_type_human(
            df['Arm ID'],
            df['Arm Name'],
            df['Study Population Description'],
            df['Arm Type'],
            df['Ethnicity*'],
            df['Race*'],
            df['Race Specify'],
            df['Description'],
            df['Sex at Birth*'],
            df['Age Event'],
            df['Subject Phenotype'],
            df['Study Location*'],
            df['Assessment Name'],
            df['Measured Behavioral or Psychological Factor*'],
            df['Measured Social Factor*'],
            df['SARS-CoV-2 Symptoms*'],
            df['Assessment_Clinical  and Demographic Data Provenance'],
            df['Assessment_Demographic Data Types Collected'],
            df['SARS-CoV2 History*'],
            df['SARS-CoV-2 Vaccine Type*'],
            df['COVID-19 Disease Severity*'],
            df['Post COVID-19 Symptoms'],
            df['COVID-19 Complications']
            )

    elif sub_section == 'Subject Type: model organism':
        df = seroFxn.edit_df(df)

        SUBJECT_ORGANISM = seroClass.subject_type_mode_organism(
            df['Arm ID'],
            df['Arm Name'],
            df['Study Population Description'],
            df['Arm Type'],
            df['Species'],
            df['Biosample Type'],
            df['Strain Characteristics'],
            df['Sex at Birth*'],
            df['Age Event'],
            df['Subject Phenotype'],
            df['Study Location*'],
            df['SARS-CoV2 History*'],
            df['SARS-CoV-2 Vaccine Type*'],
            df['COVID-19 Disease Severity*'],
            df['Post COVID-19 Symptoms'],
            df['COVID-19 Complications']
            )
    
    elif sub_section == 'planned_visit':
        df = seroFxn.edit_df(df)

        PLANNED_VISIT = seroClass.planned_visit(
            df['Visit ID'],
            df['Visit Name'],
            df['Visit Order Number'],
            df['Visit Min Start Day'],
            df['Visit Max Start Day'],
            df['Visit Start Rule']                                
    )
    
    elif sub_section == 'Study Experiment Samples':
        df = seroFxn.edit_df(df)

        EXPERIMENT_SAMPLES = seroClass.study_experiment_samples(
            df['Expt Sample User Defined ID'],
            df['Biospecimen Type*'],
            df['Biospecimen Collection Point*']                             
    )
    
    elif sub_section == 'Study Experiments':
        df = seroFxn.edit_df(df)

        STUDY_EXPERIMENTS = seroClass.study_experiment(
            df['Experiment ID'],
            df['Experiment Name'],
            df['Assay Type'],
            df['Experiment Results File Name']
    )
        
    elif sub_section == 'Reagent':
        df = seroFxn.edit_df(df)

        REAGENTS = seroClass.reagent_per_experiment(
            df['Reagent ID'],
            df['SARS-CoV-2 Antigen'],
            df['Assay Use'],
            df['Manufacturer'],
            df['Catalog #']
    )  
        
    elif sub_section == 'Results for Serology Assays':
        df = seroFxn.edit_df(df)

        RESULTS = seroClass.results(
            df['Virus Target'],
            df['Antibody Isotype'],
            df['Reporting Units'],
            df['Assay Reporting Format']
    )  
        
    else:
        print(sub_section, ': does not exist')


# In[6]:


@dataclass
class TEST:
    """1 Column to the right"""
    Study_Identifier: str = None
    Study_Name: str = None
    Publication_Title: str = None
    Study_Objective: str = None
        
    def __post_init__(self):

        self.Study_Identifier = self.Study_Identifier.replace('\n','').replace('\t','')
        self.Study_Name = self.Study_Name.replace('\n','').replace('\t','')
        self.Publication_Title = self.Publication_Title.replace('\n','').replace('\t','')

        for attr in self.fields():
            print(attr)


# In[7]:


# # Sending back to ImmPort Template #

# In[8]:


ImmPortClassNames = ['study', 'study_categorization', 'study_2_condition_or_disease',
               'arm_or_cohort','study_personnel', 'planned_visit', 'inclusion_exclusion',
               'study_2_protocol', 'study_file', 'study_link', 'study_pubmed'] 


# In[9]:


shutil.copy(PATH_basic_stdy_template, PATH_pmid_basic_stdy_template)
basic_stdy_template = load_workbook(PATH_pmid_basic_stdy_template)
bst_ws = basic_stdy_template['basic_study_design.txt']

se = seroFxn.get_sections(bst_ws, ImmPortClassNames)


# # Basic Study Template

# In[10]:


if len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type) + len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
    
    vaccine_type = list(set([i.split(';')[0] for i in SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type] + [i.split(';')[0] for i in SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type]))
    
elif len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type):
    
    vaccine_type = list(set([i.split(';')[0] for i in SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type]))
                            
elif len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
    
    vaccine_type = list(set([i.split(';')[0] for i in SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type]))
    
else:
    
    vaccine_type = []



if SUBJECT_HUMAN and SUBJECT_ORGANISM:
    print('Both human and model organism are used')
    usr_id = list(SUBJECT_HUMAN.User_Defined_ID) + list(SUBJECT_ORGANISM.User_Defined_ID)
    name = list(SUBJECT_HUMAN.Name) + list(SUBJECT_ORGANISM.Name)
    description = list(SUBJECT_HUMAN.Description) + list(SUBJECT_ORGANISM.Description)
    type_report = list(SUBJECT_HUMAN.Type_Reported) + list(SUBJECT_ORGANISM.Type_Reported)
    
    if set(SUBJECT_HUMAN.User_Defined_ID).intersection(SUBJECT_ORGANISM.User_Defined_ID):
        print("\n*** ERROR: Cannot have same User Defined ID's for AOCs:")
        print(set(SUBJECT_HUMAN.User_Defined_ID).intersection(SUBJECT_ORGANISM.User_Defined_ID))
        
if SUBJECT_HUMAN and not SUBJECT_ORGANISM:
    print('Only human is used')
    usr_id = list(SUBJECT_HUMAN.User_Defined_ID)
    name = list(SUBJECT_HUMAN.Name)
    description = list(SUBJECT_HUMAN.Description)
    type_report = list(SUBJECT_HUMAN.Type_Reported)
    vaccine_type = [i.split(';')[0] for i in list(set(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type))]
    
if SUBJECT_ORGANISM and not SUBJECT_HUMAN:
    print('Only organism is used')
    usr_id = list(SUBJECT_ORGANISM.User_Defined_ID)
    name = list(SUBJECT_ORGANISM.Name)
    description = list(SUBJECT_ORGANISM.Description)
    type_report = list(SUBJECT_ORGANISM.Type_Reported)
    vaccine_type = [i.split(';')[0] for i in list(set(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type))]


# In[12]:


# In[13]:


temp_wb = Workbook()
temp_ws = temp_wb.active

#making a temp workbook to store each section. This will be turned into df

# This is the first part of the df
for i in bst_ws.iter_rows(values_only = True,
                          max_row = se[2]+1):
    temp_ws.append(i)
    
# Starting at Arm or Cohort  



AOC = seroClass.arm_or_cohort(
    list(filter(None, usr_id)),
    list(filter(None, name)),
    list(filter(None, description)),
    list(filter(None, type_report))
)

# temp_ws.append(['User Defined ID', 'Name','Description','Type Reported'])
seroFxn.add_df(temp_ws, AOC)

seroFxn.add_df(temp_ws, STUDY_PERSONNEL)
seroFxn.add_df(temp_ws, PLANNED_VISIT)
seroFxn.add_df(temp_ws, INCLUSION_EXCLUSION)

# # Add protocol
temp_ws.append([])
temp_ws.append([PROTOCOLS.ImmPortNAME])
temp_ws.append(['Protocol ID', PROTOCOLS.Protocol_Name[1]])

seroFxn.add_df(temp_ws, STUDY_FILE)
seroFxn.add_df(temp_ws, STUDY_LINK)

seroFxn.add_df(temp_ws, STUDY_PUBMED)

# pd.DataFrame(temp_ws.values).to_csv('./TEST.csv', index=None)


# #### Filling in Study Info Sections 

# In[14]:


registryToImmportDict = pd.read_csv(registryToImmportDict_file, 
                                    header=None, 
                                    index_col=0, 
                                    squeeze=True).to_dict()

# registryToImmportDict
registryDict = {**vars(STUDY),
                **vars(STUDY_CATEGORIZATION),
                **vars(STUDY_DESIGN),
                **vars(STUDY_DETAILS),
                **vars(STUDY_PUBMED),
                **vars(COD)}

if len(''.join(vaccine_type)) != 0:
    registryDict['SARS-CoV-2_Vaccine_Type'] = vaccine_type
else:
    registryDict['SARS-CoV-2_Vaccine_Type'] = ['NA']



# Looping through ImmPort Template to get the correct order of the 'study' section
for se_number in range(se[0],se[3]):    
 
    if temp_ws["A"][se_number].value != None and registryToImmportDict.get(temp_ws["A"][se_number].value) != None:
#         print(temp_ws["A"][se_number].value)
        
        # Using a mapping key + using info in our classes to map the data
#         print(temp_ws["A"][se_number].value)
        reg_key = registryToImmportDict.get(temp_ws["A"][se_number].value).strip().replace(' ',"_").replace('*',"")
#         print(registryDict.get(reg_key))
        try:
            # If input is a list, we will turn it into a string (since it cant be a list)
            if type(registryDict.get(reg_key)) == list:
                temp_ws["B"][se_number].value = ', '.join(registryDict.get(reg_key))
            else:
                temp_ws["B"][se_number].value = registryDict.get(reg_key)

        except:
            print(f"{reg_key} did not work")
            
#     else:
#         print('blank')


# In[15]:


bsd = pd.DataFrame(temp_ws.values).replace({None: '', 'None': ''})
# bsd.to_excel(os.path.join(OUT_DIR, f'PMID{PMID}_study.xlsx'), index=False, header = False)
bsd.to_csv(os.path.join(OUT_DIR,f'{BASIC_STUDY_TEMPLATE}.txt'),
           header = False,
           index = False,
           sep = '\t')


# In[16]:


bsd.head(60)


# ### errors: None

# # Protocol 

# In[17]:


protocol_ws = load_workbook(PATH_protocols)['protocols.txt']
protocol_ws = seroFxn.remove_excess(protocol_ws)
seroFxn.add_df(protocol_ws, PROTOCOLS, add_header = False, stagger = 1)


# In[18]:


protocol_df = pd.DataFrame(protocol_ws.values).replace({None: '', 'None': ''})
protocol_df.to_csv(os.path.join(OUT_DIR,f'{PROTOCOL_TEMPLATE}.txt'),
                   header = False, 
                   index = False,
                   sep = '\t')
try:
    shutil.copy(os.path.join(BASE_DIR,'templated_data', f'{PROTOCOLS.Protocol_Name[1]}.txt'),
                os.path.join(OUT_DIR,f'{PROTOCOLS.Protocol_Name[1]}.txt'))
except:
    print(f'File: {PROTOCOLS.Protocol_Name[1]}.txt does not exist')



# # Experiment

# creating a map of the assay types to the SeroNet descriptors 
reg_description = pd.read_excel(df_path, sheet_name = map_sheet)
descriptions = dict(zip(reg_description['Unnamed: 1'][4:], reg_description['Unnamed: 2'][4:]))

#creating a df to add into the worksheet
Assay_used = STUDY_EXPERIMENTS.Experiment_Assay_Type

experiments_df = pd.DataFrame({
    'Column Name': ['']*len(Assay_used),
    'User Defined ID': [f'PMID{PMID}_exp-0'+str(i+1) for i,k in enumerate(Assay_used)],
    'Name': [Assay_used[i+1] for i, k in enumerate(Assay_used)],
    'Description': [descriptions.get(k) for i, k in enumerate(Assay_used)],
    'Measurement Technique': [STUDY_EXPERIMENTS.Experiment_Assay_Type[i+1] for i, k in enumerate(Assay_used)],
    'Study ID': [STUDY.Study_Identifier]*len(Assay_used),
    'Protocol ID(s)': [PROTOCOLS.Protocol_ID[1]]*len(Assay_used)
})


# loading experiment template and removing excess rows and columns
experiment_ws = load_workbook(PATH_experiments)['experiments.txt']
experiment_ws = seroFxn.remove_excess(experiment_ws)

# adding df to bottom of ws
seroFxn.add_df(experiment_ws, experiments_df, add_header = False)


# In[21]:


experiments_df = pd.DataFrame(experiment_ws.values).replace({None: '', 'None': ''})
experiments_df.to_csv(os.path.join(OUT_DIR,f'{EXP_TEMPLATE}.txt'),
                      header = False, 
                      index = False,
                      sep = '\t')

experiments_df



# # REAGENT

# This is creating the template even if it is filled with the defaults


if REAGENTS:
    NumReagents = len(REAGENTS.Reagent_ID)
    empty = [''] * NumReagents

    reagent_df = pd.DataFrame({
        'Column Name': empty,
        'User Defined ID': REAGENTS.Reagent_ID,
        'Name': REAGENTS.SARS_CoV_2_Antigen,
        'Description': REAGENTS.Assay_Use,
        'Manufacturer': REAGENTS.Manufacturer,
        'Catalog Number': REAGENTS.Catalog,
        'Lot Number': empty,
        'Weblink': empty,
        'Contact': empty

    })

    # loading experiment template and removing excess rows and columns
    reagent_ws = load_workbook(PATH_reagent)['reagents.Other.txt']
    reagent_ws = seroFxn.remove_excess(reagent_ws)

    # adding df to bottom of ws
    seroFxn.add_df(reagent_ws, reagent_df, add_header = False)
    reagent_df = pd.DataFrame(reagent_ws.values).replace({None: '', 'None': ''})

    # saving df
    reagent_df.to_csv(os.path.join(OUT_DIR,f'{REAGENT_TEMPLATE}.txt'),
                       header = False, 
                       index = False,
                       sep = '\t')

    reagent_df


# In[23]:


reagent_df


"""
# # SUBJECT: HUMAN

# - if human, use human
# - if animal, use animal sheet 
# - use length as a predictor

HUMAN SHOULD BE FOR HUMAN CELLS LINES. DO NOT PUT IN ORGANISM

"""


if SUBJECT_HUMAN:
    species = SUBJECT_HUMAN.User_Defined_ID
    empty = ['']*len(species)
    
    if len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type): #splitting correctly
        try: #check if there are multiple vaccine names in this section
            vaccine_type = []
            vaccine_name = []

            for vaccines_per_sub in SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type:
                v_type = []
                v_name = []

                for each in vaccines_per_sub.split('|'):
                    v_type.append(each.split(";")[1].strip())
                    v_name.append(each.split(";")[0].strip())

                vaccine_type.append(' | '.join(v_type))
                vaccine_name.append(' | '.join(v_name))
        except:
            vaccine_name = [i.split('; ')[1] for i in vaccines]
            vaccine_type = [i.split('; ')[0] for i in vaccines]
        
    else:
        vaccine_name = empty
        vaccine_type = empty
        
    
    SUBJECT_human_df = pd.DataFrame({
        'Column Name': empty,
        'Subject ID': [f"PMID{PMID}_human_subject-0{int(i+1)}" for i in range(len(species))],
        'Arm Or Cohort ID': SUBJECT_HUMAN.User_Defined_ID, #I feel like this needs to be defined
        'Gender': SUBJECT_HUMAN.Sex_at_Birth, 
        'Min Subject Age': 0, # add this to validator in dataclass 
        'Max Subject Age': 89, # add this to validator in dataclass 
        'Age Unit': [STUDY_DETAILS.Age_Unit]*len(species),
        'Age Event': SUBJECT_HUMAN.Age_Event, 
        'Age Event Specify': empty,
        'Subject Phenotype': empty, 
        'Subject Location': SUBJECT_HUMAN.Study_Location,
        'Ethnicity': SUBJECT_HUMAN.Ethnicity,
        'Race': SUBJECT_HUMAN.Race, 
        'Race Specify': empty,
        'Description': empty, 
        'Result Separator Column': empty, 
        'Exposure Process Reported': ['unknown']*len(species) , #not sure 
        'Exposure Material Reported': vaccine_name, # change this
        'Exposure Material ID': vaccine_type,
        'Disease Reported': empty,
        'Disease Ontology ID': empty,
        'Disease Stage Reported': empty 
    })
    
    # loading experiment template and removing excess rows and columns
    human_ws = load_workbook(PATH_subject_human)['subjectHumans.txt']
    human_ws = seroFxn.remove_excess(human_ws)

    # adding df to bottom of ws
    seroFxn.add_df(human_ws, SUBJECT_human_df, add_header = False)
    SUBJECT_human_df = pd.DataFrame(human_ws.values).replace({None: '', 'None': ''})

#     # saving df
    SUBJECT_human_df.to_csv(os.path.join(OUT_DIR,f'{SUBJ_HUMAN_TEMPLATE}.txt'),
                       header = False, 
                       index = False,
                       sep = '\t')

#     print(SUBJECT_human_df)
    print ("SUBJECT: human data created")
# else:
#     print('no human data')
    



"""
# SUBJECT: ORGANISM

For Syrain Hamster: https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?name=Syrian+hamsters 

Use latin name -- or look at lk species
https://www.immport.org/shared/templateDocumentation?tab=2&table=lk_species


"""


if len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
    for i,k in enumerate(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
        print(i)
        if k.strip().lower() not in VARS_TO_CLEAN and k != 'n/a': 
            try:
                vaccine_name.append(k.split('; ')[0])
                vaccine_type.append(k.split('; ')[1])
            except:
                vaccine_name.append(k)
                vaccine_type.append('')

        else:
            vaccine_name = empty
            vaccine_type = empty
else:
            vaccine_name = empty
            vaccine_type = empty




if SUBJECT_ORGANISM:  # Not sure how this plays out. Might need to do a mock one. Will it be 1 subject per study?
    species = SUBJECT_ORGANISM.User_Defined_ID
    empty = ['']*len(species)
    vaccine_name = []
    vaccine_type = []
    
    if len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
        for i,k in enumerate(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
            if k.strip().lower() not in VARS_TO_CLEAN and k != 'n/a': 
                try:
                    vaccine_name.append(k.split('; ')[0])
                    vaccine_type.append(k.split('; ')[1])
                except:
                    vaccine_name.append(k)
                    vaccine_type.append('')
                    
            else:
                vaccine_name = empty
                vaccine_type = empty

    else:
            vaccine_name = empty
            vaccine_type = empty

    # print ("SUBJECT_organism data")
    SUBJECT_organism_df = pd.DataFrame({
        'Column Name':empty,
        'Subject ID': [f"PMID{PMID}_organism_subject-0{int(i+1)}" for i in range(len(species))], #SUBJECT_ORGANISM.User_Defined_ID, #
        'Arm Or Cohort ID': SUBJECT_ORGANISM.User_Defined_ID, #I feel like this needs to be defined
        'Gender': SUBJECT_ORGANISM.Sex_at_Birth,
        'Min Subject Age': [STUDY_DETAILS.Minimum_Age]*len(species),
        'Max Subject Age': [STUDY_DETAILS.Maximum_Age]*len(species),
        'Age Unit': [STUDY_DETAILS.Age_Unit] * len(species),
        'Age Event': SUBJECT_ORGANISM.Age_Event,
        'Age Event Specify': empty,
        'Subject Phenotype': empty, ## This does not exist
        'Subject Location': SUBJECT_ORGANISM.Study_Location, # This needs to be changed. SUBJECT.Study_Location,
        'Species': SUBJECT_ORGANISM.Species,
        'Strain': SUBJECT_ORGANISM.Biosample_Types,
        'Strain Characteristics': SUBJECT_ORGANISM.Strain_Characteristics,
        'Result Separator Column': empty,
        'Exposure Process Reported': ['unknown']*len(species),
        'Exposure Material Reported': vaccine_name,
        'Exposure Material ID': vaccine_type, 
        'Disease Reported': [COD.Reported_Health_Condition[0]]*len(species), # also maybe not...
        'Disease Ontology ID': empty,
        'Disease Stage Reported': SUBJECT_ORGANISM.COVID_19_Disease_Severity
    })
    
        # loading experiment template and removing excess rows and columns
    organism_ws = load_workbook(PATH_subject_organism)['subjectAnimals.txt']
    organism_ws = seroFxn.remove_excess(organism_ws)

    # adding df to bottom of ws
    seroFxn.add_df(organism_ws, SUBJECT_organism_df, add_header = False)
    SUBJECT_organism_df = pd.DataFrame(organism_ws.values).replace({None: '', 'None': ''})

    # saving df
    SUBJECT_organism_df.to_csv(os.path.join(OUT_DIR,f'{SUBJ_ORGANISM_TEMPLATE}.txt'),
                       header = False, 
                       index = False,
                       sep = '\t')

    SUBJECT_organism_df
    
#     subjectAnimals



# # ASSESSMENT 
# There will be 1 excel document produced per assesment. (wide format) <br>
#     - 1 subj with multiple panels -> seperate 
#     - 1 panel with multiple subject -> 1 row per, long format
# assesment doc. need to be uploaded one at a time
# 

# In[55]:


if SUBJECT_HUMAN:
    print("Assays Used")
    allCovSymtoms = list(SUBJECT_HUMAN.Measured_Behavioral_or_Psychological_Factor) + list(SUBJECT_HUMAN.Measured_Social_Factor) + list(SUBJECT_HUMAN.SARS_CoV_2_Symptoms)
 
    NumAssessments = len(SUBJECT_HUMAN.User_Defined_ID)
    assessmet_type = []
    status = []

    for symptom in list(set(allCovSymtoms)):
        if symptom:
            assessment_name = symptom

            for i in range(NumAssessments): ### THIS NEEDS TO BE 1 cell to the right. So the actual vaules 

                i += 1
                if SUBJECT_HUMAN.Measured_Behavioral_or_Psychological_Factor[i] == assessment_name:
                    # 'Measured Behavioral or Psychological Factor'
                    assessmet_type.append(SUBJECT_HUMAN.Measured_Behavioral_or_Psychological_Factor[i])
                    status.append(SUBJECT_HUMAN.Assessment_Demographic_Data_Types_Collected[i])
                elif SUBJECT_HUMAN.Measured_Social_Factor[i] == assessment_name:
                    # 'Measured Social Factor'
                    assessmet_type.append(SUBJECT_HUMAN.Measured_Social_Factor[i])
                    status.append(SUBJECT_HUMAN.Assessment_Demographic_Data_Types_Collected[i])
                elif SUBJECT_HUMAN.SARS_CoV_2_Symptoms[i] == assessment_name:
                    # 'SARS-CoV-2 Symptoms'
                    assessmet_type.append(SUBJECT_HUMAN.SARS_CoV_2_Symptoms[i])
                    status.append(SUBJECT_HUMAN.Assessment_Demographic_Data_Types_Collected[i])
                else:
                    print('No assessment found')

            empty = ['']*len(assessmet_type) 
            # need to figure out named report 
            # not sure where the data goes for the assessments

            assessment_df = pd.DataFrame({
                'Column Name': empty,
                'Subject ID': [f"PMID{PMID}subject-0{int(i+1)}" for i in range(len(assessmet_type))],  
                'Assessment Panel ID': [f"PMID{PMID}assessment_{assessment_name}-0{int(i+1)}" for i in range(len(assessmet_type))],
                'Study ID': [STUDY.Study_Identifier]*len(assessmet_type),
                'Name Reported': assessment_name*len(assessmet_type),
                'Assessment Type': assessmet_type,
                'Status': status,
                'CRF File Names': empty,
                'Result Separator Column': empty,
                'User Defined ID': [f"PMID{PMID}assessment-0{int(i+1)}" for i in range(len(assessmet_type))], #
                'Planned Visit ID': empty, # IM going to make this empty for now PLANNED_VISIT.User_Defined_ID[i],
                'Name Reported ': [f"component_{assessment_name}"]*len(assessmet_type), # space might not work
                'Study Day': i-1, # Not sure, we dont capture study day 
                'Age At Onset Reported': empty,
                'Age At Onset Unit Reported': empty,
                'Is Clinically Significant': empty,
                'Location Of Finding Reported': empty,
                'Organ Or Body System Reported': empty,
                'Result Value Reported': 'NA',
                'Result Unit Reported': empty,
                'Result Value Category': empty,
                'Subject Position Reported': empty,
                'Time Of Day': empty,
                'Verbatim Question': empty,
                'Who Is Assessed': empty
            })

            # loading experiment template and removing excess rows and columns
            assessment_ws = load_workbook(PATH_assessment)['assessments.txt']
            assessment_ws = seroFxn.remove_excess(assessment_ws)

            # adding df to bottom of ws
            seroFxn.add_df(assessment_ws, assessment_df, add_header = False)
            assessment_df = pd.DataFrame(assessment_ws.values).replace({None: '', 'None': ''})
        #     assessment_df.rename(columns=lambda x: x.strip())
    #         print(assessment_df)
                # saving file

            if len(assessment_name) < 7:
                outname = assessment_name
            if len(assessment_name) > 44:
                outname = "combined"
            else:
                outname = assessment_name[:5]

            assessment_df.to_csv(os.path.join(OUT_DIR,f'panel_{outname}_{ASSESSMENT_TEMPLATE}.txt'),
                               header = False, 
                               index = False,
                               sep = '\t')
            # #     print(assessment_df.head())
else:
    print("No Human Subjects: assessments not recorded")


# In[57]:
### POST: copy log to folder

CD = os.getcwd()
today = dt.datetime.today().strftime('%Y_%m_%d')

try:
    os.mkdir(os.path.join(CD,BASE_DIR,"log"))
    print(f'Creating log')
except FileExistsError:
    print('Will not create log - already exists')
    pass

shutil.copyfile(os.path.join(CD,"log",f"Registry_{today}.log"), os.path.join(CD,BASE_DIR,"log",f"Registry_{today}.log"))

