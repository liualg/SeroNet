import pandas as pd
import json
import re
import numpy as np
from datetime import datetime

import glob
from fuzzywuzzy import fuzz
from openpyxl import load_workbook

import datetime as dt
import logging
import os
import sys
'''
v1.4
- Updated Logs
- Yes / Include was changed -- please checkdr
'''
CD = os.getcwd()

if not os.path.exists(os.path.join(CD, "log")):
    os.mkdir(os.path.join(CD, "log"))

today = dt.datetime.today().strftime('%Y_%m_%d')
logging.basicConfig(filename=os.path.join(CD, "log", f"Registry_{today}.log"), level=logging.DEBUG,
                    format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p') #filemode='w+', 

STUDY_PUBMED = 13

STUDY_IDENTIFIER = 16
STUDY_NAME = 17
PUBLICATION_TITLE = 18
STUDY_OBJECTIVE = 19
STUDY_DESCRIPTION = 20
PRIMARY_INSTITUTION_NAME = 21

STUDY_PERSONNEL = 25
STUDY_FILE = 39
STUDY_LINK = 45
STUDY_CATEGORIZATION = 49
RESEARCH_FOCUS = 50
STUDY_TYPE = 51
KEYWORDS = 52
STUDY_DESIGN = 55
CLINICAL_STUDY_DESIGN = 56
IN_SILICO_MODEL_TYPE = 57
PROTOCOLS = 60
PROTOCOL_ID = 61
PROTOCOL_FILE_NAME = 62
PROTOCOL_NAME = 63
PROTOCOL_DESCRIPTION = 64
PROTOCOL_TYPE = 65
CONDITION_OR_DISEASE = 68
REPORTED_HEALTH_CONDITION = 69

VIRUS_VARIANT = 71
SARS_SARS_COV_2_VARIANT = 72

INTERVENTION_AGENT = 74
SARS_COV_2_VACCINE_TYPE = 75
STUDY_DETAILS = 78
CLINICAL_OUTCOME_MEASURE = 79
ENROLLMENT_START_DATE = 80
ENROLLMENT_END_DATE = 81
NUMBER_OF_STUDY_SUBJECTS = 82
AGE_UNIT = 83
MINIMUM_AGE = 84
MAXIMUM_AGE = 85

INCLUSION_EXCLUSION = 88
INCLUSION_ID = 89
INCLUSION_CRITERION = 90
INCLUSION_CRITERION_CATEGORY = 91

HUMAN_ARM_ID = 95
HUMAN_ARM_NAME = 96
HUMAN_STUDY_POPULATION_DESCRIPTION = 97
HUMAN_ARM_TYPE = 98
HUMAN_ETHNICITY = 99
HUMAN_RACE = 100
HUMAN_RACE_SPECIFY = 101
HUMAN_DESCRIPTION = 102
HUMAN_SEX_AT_BIRTH = 103
HUMAN_AGE_EVENT = 104
HUMAN_SUBJECT_PHENOTYPE = 105
HUMAN_STUDY_LOCATION = 106
HUMAN_ASSESSMENT_NAME = 107
HUMAN_MEASURED_BEHAVIORAL_OR_PYSCHOLOGICAL_FACTOR = 108
HUMAN_MEASURED_SOCIAL_FACTOR = 109
HUMAN_SARS_COV_2_SYMPTOMS = 110
HUMAN_ASSESSMENT_CLINICAL_AND_DEMOGRAPHIC_DATA_PROVENANCE = 111
HUMAN_ASSESSMENT_DEMOGRAPHIC_DATA_TYPES_COLLECTED = 112
HUMAN_SARS_COV_2_HISTORY = 113
HUMAN_SARS_COV_2_VACCINE_TYPE = 114
HUMAN_COVID_19_DISEASE_SEVERITY = 115
HUMAN_POST_COVID_19_SYMPTOMS = 116
HUMAN_COVID_19_COMPLICATIONS = 117

MODEL_ARM_ID = 121
MODEL_ARM_NAME = 122
MODEL_STUDY_POPULATION_DESCRIPTION = 123
MODEL_ARM_TYPE = 124
MODEL_SPECIES = 125
MODEL_BIOSAMPLE_TYPE = 126
MODEL_STRAIN_CHARACTERISTICS = 127
MODEL_SEX_AT_BIRTH = 128
MODEL_AGE_EVENT = 129
MODEL_SUBJECT_PHENOTYPE = 130
MODEL_STUDY_LOCATION = 131
MODEL_SARS_COV_2_HISTORY = 132
MODEL_SARS_COV_2_VACCINE_TYPE = 133
MODEL_COVID_19_DISEASE_SEVERITY = 134
MODEL_POST_COVID_19_SYMPTOMS = 135
MODEL_COVID_19_COMPLICATIONS = 136

PLANNED_VISIT = 139
VISIT_ID = 140
VISIT_NAME = 141
VISIT_ORDER_NUMBER = 142
VISIT_MIN_START_DAY = 143
VISIT_MAX_START_DAY = 144
VISIT_START_RULE = 145

EXPERIMENTS = 149
ASSOCIATED_ARM_IDS = 150
ASSOCIATED_FIRST_PLANNED_VISIT_ID = 151
ASSAY_TYPE = 152
EXPERIMENT_NAME = 153
EXPERIMENT_RESULTS_FILE_NAME = 154
BIOSPECIMEN_TYPE = 155
BIOSPECIMEN_COLLECTION_POINT = 156
SARS_COV_2_ANTIGEN = 157
ASSAY_USE = 158
MANUFACTURER = 159
CATALOG_NUMBER = 160
VIRUS_TARGET = 161
ANTIBODY_ISOTYPE = 162
REPORTING_UNITS = 163
ASSAY_REPORTING_FORMAT = 164

STATUS_NOTE = 167

## loading workbooks and data

json_seronet_dict = pd.read_excel("./dictionary/Facetdict.xlsx",
                                 header=None)
json_seronet_dict = dict(
    zip(json_seronet_dict.loc[:][0],json_seronet_dict.loc[:][1])
)

## creating dictionary of correct spelling from most recent template
try:
    try:
        wb = load_workbook(
            glob.glob("/Users/liualg/Library/CloudStorage/OneDrive-SharedLibraries-NationalInstitutesofHealth/NCI-FNL SeroNet Team - Curation channel/*.xlsm")[0],
                          )
    except:
        wb = load_workbook(
            glob.glob("/Users/liualg/Library/CloudStorage/OneDrive-NationalInstitutesofHealth/Curation channel/*.xlsm")[0],
                          )
except:
    sys.exit("Workbook not loaded correctly") 

sheet = wb['Cntrl\'d Vocab']
immport_dict = dict()

for column in sheet.iter_cols( #sheet.max_column, 
                              max_row=sheet.max_row, 
                              min_row=4):
    name = []
    start_index = []
    for icol, cell in enumerate(column):
        if icol == 0:
            name.append(cell.value)
            start_index.append(icol)
        else:
            if cell.value is not None and (column[icol-1].value is None):
                name.append(cell.value)
                start_index.append(icol)
    
    try:
        for i in range(len(start_index)-1):
            # This does the main rows
            immport_dict[name[i]] = [k.value for k in column[start_index[i]+1 : start_index[i+1]] if k.value is not None]
        #this does the last row
        immport_dict[name[-1]] = [k.value for k in column[start_index[i+1]+1 :] if k.value is not None]
    
    except:
        #this does the first row 
        immport_dict[name[-1]] = [k.value for k in column[start_index[i-1]+1 :] if k.value is not None]
        pass


# I want a dict per study to memorize the choices i made
# it to be the table name
# table_name, the word im checking, the outcome from last time
df_check = pd.DataFrame(columns=['field','lookup_word','user_input'])

def get_closest_lookup(word, lookup_table, table_name):
    global df_check
    correct_word = ''
    closest_lookup = ''
    check_list = []
    
    if word in ['', 'N/A', 'n/a', 'N/a', 'n/A', np.nan, None]:
        print(f'[NOTE]:: Field empty: {table_name}')
        closest_lookup = 'Not Applicable' #LIU 11023
    
    elif word in lookup_table:
        closest_lookup = word

    else:
        matching = [s for s in lookup_table if word.lower() in s.lower()]

        if len(matching) == 1:
            closest_lookup = matching[0]
            print(f"INFO:: [Fuzzy] Closest Match Found. Changing {word} => {closest_lookup}")
            logging.info(f"INFO:: [Fuzzy] Closest Match Found. Changing {word} => {closest_lookup}")

        
        else:
            if len(matching) > 1:
                lookup_table = matching
            else:
                pass 

            for potential in lookup_table:
                check_list.append((potential, 
                                      fuzz.ratio(word, potential))
                                     )
            
            # if the list 
            if len(check_list) == 1:
                print(f"[INFO]:: [Fuzzy] {word} => {check_list[0][0]} : score {check_list[0][1]}")
                logging.info(f"[INFO]:: [Fuzzy] {word} => {check_list[0][0]} : score {check_list[0][1]}")
                closest_lookup = check_list[0][0]


            elif len(check_list) > 1:
                check_list.sort(key=lambda a: a[1], reverse=True)
                previous_inputs = df_check['user_input'][(df_check['field']==table_name) & (df_check['lookup_word'] == word)]

                if float(check_list[0][1])/np.std([n[1] for n in check_list]) >= 3 and float(check_list[0][1]) > 75: #checking if i cann bypass the score choice (stat. sig.)
                # if float(check_list[0][1])/np.std([n[1] for n in check_list[:3]]) >= 3: #checking if i cann bypass the score choice (stat. sig.)

                    print(f"[INFO]:: [Fuzzy] {word} => {check_list[0][0]} : score {check_list[0][1]}")
                    logging.info(f"[INFO]:: [Fuzzy] {word} => {check_list[0][0]} : score {check_list[0][1]}")
                    closest_lookup = check_list[0][0]

                # checking to see if work exists in the choice table already
                elif len(previous_inputs) > 0:
                    closest_lookup = previous_inputs[0]
                    print(f"Previous Input: {table_name}, {word} => {previous_inputs[0]}")
                    logging.info(f"Previous Input: {table_name}, {word} => {previous_inputs[0]}")

                else:
                    
                    print(f'\n\n######  ACTION REQUIRED ###### \n[{table_name}] Designate replacement word for: {word}')
                    print('Top three choices:')
                    for top3 in check_list[:3]:
                        print(top3)
                    # print(check_list)

                    user_resp = input(f'[{table_name}] Designate replacement word for: {word}: ')
                    try:
                        user_resp = int(user_resp)
                    except:

                        while not isinstance(user_resp, int):
                            print('\n*** Please use a number to designate word you want substituted ***\n\t-OR type more or more options\n\t-OR type exit')
                            try:
                                user_resp = input(f'[{table_name}] Designate replacement word for {word}: ')
                                user_resp = int(user_resp)
                            except:
                                user_resp = str(user_resp)
                            
                            if user_resp == 'more':
                                for ind, top3 in enumerate(check_list):
                                    print(f"{ind+1}. {top3}")
                            if user_resp == 'exit':
                                break

                    closest_lookup = check_list[int(user_resp)-1][0]
                    logging.info(f"INFO:: [User Change] {word} => {closest_lookup}")
                    df_check = pd.concat([df_check, pd.DataFrame({'field':[table_name],'lookup_word':[word],'user_input':[closest_lookup]})])

            else:
                pass

    return closest_lookup

def check_spelling(word_in_questions, facet_to_check):
    lk_table = immport_dict.get(json_seronet_dict.get(facet_to_check))

    if isinstance(word_in_questions,str):
        if word_in_questions:
            correct_word = get_closest_lookup(word_in_questions,lk_table,json_seronet_dict.get(facet_to_check))
        else:
            correct_word = ''

    elif isinstance(word_in_questions, list):
        correct_word = []
        if not all(('' == item or None is item) for item in word_in_questions): #must be this?
            for words in word_in_questions:
                correct_word.append(get_closest_lookup(words,lk_table,json_seronet_dict.get(facet_to_check)))
        else:
            pass
    else:
        print('something is wrong')
        pass

    return correct_word
#
# Remove characters or escape characters that corrupt JSON
#
def cleanData(s):
    """Removes characters in the input string that will corrupt the final JSON object"""
    if pd.isna(s):
        return ""
    else:
        if isinstance(s, str):
            r1 = re.compile("\n|\t|\r")
            r2 = re.compile('"')
            s = r1.sub(" ", s)
            s = r2.sub('\\"', s)
            s = s.strip()
        elif isinstance(s, datetime):
            s = s.strftime('%Y-%m-%d')
        else:
            pass
        return s

def replace_na(cleaned_data):
    if pd.isna(cleaned_data) or cleaned_data in ['', 'N/A', 'n/a', np.nan, None]:
        return 'Not Applicable'
    else:
        return cleaned_data 


def parse_sv(df, line_number, index):
    return list(df.loc[line_number])[index]


def parse_clean_sv(df, line_number, index):
    """Cleans and returns data in a List"""
    return cleanData(list(df.loc[line_number])[index])


def parse_clean_mv_split(df, line_number, index):
    """Splits line using th "|" character and returns a List of clean data"""
    if pd.isna(df.loc[line_number][index]):
        return []
    else:

        return [cleanData(item.strip()) for item in re.split("\|", df.loc[line_number][index].replace(" I ", "|").replace(",","|"))]


def parse_clean_mv_split_I(df, line_number, index):
    """Splits line using th "I" character and returns a List of clean data"""
    return [cleanData(item) for item in re.split(" I ", df.loc[line_number][index])]


def parse_registry_template(df, template):
    """Controls the order of extracting data from the Excel sheet

       Parameters:
           df (dataframe): Panda's dataframe
           template (dictionary): Dictionary to hold parse data

        Returns:
            None
    """
    
    parse_pubmed(df, template)
    parse_study(df, template)
    parse_study_personnel(df, template)
    parse_study_file(df, template)
    parse_study_link(df, template)
    parse_study_categorization(df, template)
    parse_study_design(df, template)
    parse_protocol(df, template)
    parse_condition_or_disease(df, template)
    parse_variant(df, template)
    parse_intervention(df, template)
    parse_study_details(df, template)
    parse_subject_human(df, template)
    parse_model_organism(df, template)
    parse_planned_visit(df, template)
    parse_experiment(df, template)
    parse_inclusion_exclusion(df, template)
    parse_status_note(df, template)


def parse_pubmed(df, template):
    """Parse the single valued pubmed_id"""

    template['pubmed_id'] = parse_sv(df, STUDY_PUBMED, 2)


def parse_study(df, template):
    """Parse top level single valued study information"""

    template['study_identifier'] = parse_clean_sv(df, STUDY_IDENTIFIER, 2)
    template['study_name'] = parse_clean_sv(df, STUDY_NAME, 2)
    template['publication_title'] = parse_clean_sv(df, PUBLICATION_TITLE, 2)
    template['study_objective'] = parse_clean_sv(df, STUDY_OBJECTIVE, 2)
    template['study_description'] = parse_clean_sv(df, STUDY_DESCRIPTION, 2)
    template['primary_institution_name'] = parse_clean_sv(df, PRIMARY_INSTITUTION_NAME, 2)


def parse_study_personnel(df, template):
    """Parse the study personnel information
    
       There may be serveral study personnel with one person in each column
       spanning multiple lines in the spread sheet.
    """

    studyPersonnel = []
    personnel = list(df.loc[STUDY_PERSONNEL])
    for idx, val in enumerate(personnel[2:], start=2):
        if not pd.isna(val):
            obj = {
                "personnel_id": val,
                "honorific": parse_clean_sv(df, STUDY_PERSONNEL + 1, idx),
                "last_name": parse_clean_sv(df, STUDY_PERSONNEL + 2, idx),
                "first_name": parse_clean_sv(df, STUDY_PERSONNEL + 3, idx),
                "suffixes": parse_clean_sv(df, STUDY_PERSONNEL + 4, idx),
                "organization": parse_clean_sv(df, STUDY_PERSONNEL + 5, idx),
                "orchid_id": parse_clean_sv(df, STUDY_PERSONNEL + 6, idx),
                "email": parse_clean_sv(df, STUDY_PERSONNEL + 7, idx),
                "seronet_title_in_study":  parse_clean_sv(df, STUDY_PERSONNEL + 8, idx),
                "role_in_study": parse_clean_sv(df, STUDY_PERSONNEL + 9, idx),
                "site_name": parse_clean_sv(df, STUDY_PERSONNEL + 10, idx)
            }
            studyPersonnel.append(obj)
        else:
            pass
    
    template['study_personnel'] = studyPersonnel


def parse_study_file(df, template):
    """Parse the study file information
    
       There may be serveral study files with one study file in each column
       spanning multiple lines in the spread sheet.
    """

    studyFile = []
    study_files = list(df.loc[STUDY_FILE])
    for idx, val in enumerate(study_files[2:], start=2): 
        if not pd.isna(val):
            obj = {
               "study_file_name": val,
               "study_file_description": parse_clean_sv(df, STUDY_FILE + 1, idx),
               "study_file_type": parse_clean_sv(df, STUDY_FILE + 2, idx)
            }
            studyFile.append(obj)
        else:
            pass
    template['study_file'] = studyFile

    
def parse_study_link(df, template):
    """Parse the study link information
    
       There may be serveral study links with one study link in each column
       spanning multiple lines in the spread sheet.
    """
    studyLink = []
    study_links = list(df.loc[STUDY_LINK])
    for idx, val in enumerate(study_links[2:], start=2): 
        if not pd.isna(val):
            obj = {
               "link_name": val,
               "value": parse_clean_sv(df, STUDY_LINK +1, idx)
            }
            studyLink.append(obj)
        else:
            pass
    template['study_link'] = studyLink


def parse_study_categorization(df, template):
    """ Parse the Study Categorization section"""
    template['research_focus'] = check_spelling(parse_clean_sv(df, RESEARCH_FOCUS, 2), 'research_focus')
    template['study_type'] = replace_na(check_spelling(parse_clean_sv(df, STUDY_TYPE, 2), 'study_type'))

    # Keywords 
    keyword = []
    keywords = (list(df.loc[KEYWORDS])[2]).replace(";"," | ").replace(" I ", "|").replace(",","|").split("|")
    for k in keywords:
        keyword.append(k.strip())
    template['keyword'] = keyword


def parse_study_design(df, template):
    """Parse the Study Design section"""

    template['clinical_study_design'] = replace_na(check_spelling(replace_na(parse_clean_sv(df, CLINICAL_STUDY_DESIGN, 2)), 'clinical_study_design'))
    template['in_silico_model_type'] = replace_na(check_spelling(parse_clean_sv(df, IN_SILICO_MODEL_TYPE, 2),'in_silico_model_type'))


def parse_protocol(df, template):
    """Parse the Protocol section

       There may be serveral protocols with one protocol in each column
       spanning multiple lines in the spread sheet.
    """

    protocol = {}
    protocol['protocol_id'] = parse_clean_sv(df, PROTOCOL_ID, 2)
    protocol['protocol_file_name'] = parse_clean_sv(df, PROTOCOL_FILE_NAME, 2)
    protocol['protocol_name'] = parse_clean_sv(df, PROTOCOL_NAME, 2)
    protocol['protocol_description'] = parse_clean_sv(df, PROTOCOL_DESCRIPTION, 2)
    protocol['protocol_type'] = parse_clean_sv(df, PROTOCOL_TYPE, 2)

    template['protocol'] = protocol


def parse_condition_or_disease(df, template):
    """Parse the Condition or Disease line
        
        This line can contain one or more condtions current separated using "I".
    """
    rhc = list(df.loc[REPORTED_HEALTH_CONDITION])[2].replace("|"," I ").replace(",", " I ")
    values = re.split(" I ", rhc)
    reported_health_condition = []
    for c in values:
        #liu3
        reported_health_condition.append(check_spelling(cleanData(c),'reported_health_condition'))

    template['reported_health_condition'] = reported_health_condition

def parse_variant(df, template):
    """Parse the Intervention Agent line

        This line can contain one or more agents separated using "|"
    """

    values = parse_clean_mv_split(df, SARS_SARS_COV_2_VARIANT, 2)
    variant_types = []
    for v in values:
        variant_types.append(cleanData(v))

    template['sars_cov_2_variant'] = list(set(check_spelling(variant_types, 'virus_target')))

def parse_intervention(df, template):
    """Parse the Intervention Agent line

        This line can contain one or more agents separated using "|"
    """

    values = parse_clean_mv_split(df, SARS_COV_2_VACCINE_TYPE, 2)
    vaccine_types = []
    for v in values:
        vaccine_types.append(cleanData(v))

    template['sars_cov_2_vaccine_type'] = check_spelling(vaccine_types, 'sars_cov_2_vaccine_type-study')


def parse_study_details(df, template):
    """Parse the Study Detail seciont
        
        Each line contains one single value.
    """

    template['clinical_outcome_measure'] = replace_na(parse_clean_sv(df, CLINICAL_OUTCOME_MEASURE, 2))
    template['enrollment_start_date'] = parse_clean_sv(df, ENROLLMENT_START_DATE, 2)
    template['enrollment_end_date'] = parse_clean_sv(df, ENROLLMENT_END_DATE, 2)
    template['number_of_study_subjects'] = parse_clean_sv(df, NUMBER_OF_STUDY_SUBJECTS, 2)
    template['age_unit'] = parse_clean_sv(df, AGE_UNIT, 2)
    template['minimum_age'] = parse_clean_sv(df, MINIMUM_AGE, 2)
    template['maximum_age'] = parse_clean_sv(df, MAXIMUM_AGE, 2)


def parse_inclusion_exclusion(df, template):
    """Parse the inclusion/exclusion section

       Some of the entries in this section are being treated as special cases,
       because they eventually need to be treated as Yes/No facets in the CDT.
    """
    inclusion_exclusion = []
    ids = list(df.loc[INCLUSION_ID])
    
    #
    # Set the following properties to "Not Specified"
    # If they are in the template they will be updated to "Yes" or "No"
    #
    template['geriatric_subjects'] = "Not Specified"
    template['pediatric_subjects'] = "Not Specified"
    template['pregnant_subjects'] = "Not Specified"
    template['sars_cov_2_antibodies_measured'] = "Not Specified"
    template['survey_instrument_shared'] = "Not Specified"
    template['who_disease_severity_scale_used'] = "Not Specified"

    for idx, val in enumerate(ids[2:], start=2):
        if not pd.isna(val):
            obj = {
               "inclusion_exculusion_id": val,
               "inclusion_criterion": parse_clean_sv(df, INCLUSION_CRITERION, idx),
               "inclusion_criterion_category": parse_clean_sv(df, INCLUSION_CRITERION_CATEGORY, idx)
            }


            inclusion_exclusion_crit = [
            'Geriatric subjects', 'Pediatric subjects','SARS-CoV-2 Antibodies Measured',
            'Survey Instrument shared','WHO disease severity scale used'
            ]

            def check_ic_list(inputs1, obj1, template1):
                for i in inputs1:
                    i_lower = i.lower().replace(' ','_').replace('-','_')

                    if  obj1['inclusion_criterion'] == i:
                        if obj['inclusion_criterion_category'].lower() == 'inclusion':
                            template1[i_lower] = "Yes"

                        elif obj1['inclusion_criterion_category'].lower() == 'yes':
                            template1[i_lower] = "Yes"
                            obj1['inclusion_criterion_category'] = 'Inclusion'

                        elif obj1['inclusion_criterion_category'].lower() == 'no':
                            template1[i_lower] = "No"
                            obj1['inclusion_criterion_category'] = 'Exclusion'

                        else:
                            template1[i_lower] = "No"
                    else:
                        if obj1['inclusion_criterion_category'].lower() == 'yes':
                            obj1['inclusion_criterion_category'] = 'Inclusion'

                        elif obj1['inclusion_criterion_category'].lower() == 'no':
                            obj1['inclusion_criterion_category'] = 'Exclusion'
                        else:
                            pass

                return obj1


            obj = check_ic_list(inclusion_exclusion_crit,obj, template)

            inclusion_exclusion.append(obj)
            
            # include_cri = ["inclusion", "yes"]
            
            # if  obj['inclusion_criterion'] == "Geriatric subjects":
            #     if obj['inclusion_criterion_category'].lower() in include_cri :
            #         template['geriatric_subjects'] = "Yes"
            #     else:
            #         template['geriatric_subjects'] = "No"

            # if  obj['inclusion_criterion'] == "Pediatric subjects":
            #     if obj['inclusion_criterion_category'].lower() in include_cri:
            #         template['pediatric_subjects'] = "Yes"
            #     else:
            #         template['pediatric_subjects'] = "No"

            # if  obj['inclusion_criterion'] == "Pregnant subjects":
            #     if obj['inclusion_criterion_category'].lower() in include_cri:
            #         template['pregnant_subjects'] = "Yes"
            #     else:
            #         template['pregnant_subjects'] = "No"

            # if  obj['inclusion_criterion'] == "SARS-CoV-2 Antibodies Measured":
            #     if obj['inclusion_criterion_category'].lower()  in include_cri:
            #         template['sars_cov_2_antibodies_measured'] = "Yes"
            #     else:
            #         template['sars_cov_2_antibodies_measured'] = "No"

            # if  obj['inclusion_criterion'] == "Survey Instrument shared":
            #     if obj['inclusion_criterion_category'].lower() in include_cri:
            #         template['survey_instrument_shared'] = "Yes"
            #     else:
            #         template['survey_instrument_shared'] = "No"

            # if  obj['inclusion_criterion'] == "WHO disease severity scale used":
            #     if obj['inclusion_criterion_category'].lower() in include_cri:
            #         template['who_disease_severity_scale_used'] = "Yes"
            #     else:
            #         template['who_disease_severity_scale_used'] = "No"

        else:
            pass
    
    template['inclusion_exclusion'] = inclusion_exclusion


def parse_subject_human(df, template):
    """Parse the Subject Type: human section

       Some rows in this section contain single value and some contain multi-valued content

       There may be serveral entries with one entry in each column
       spanning multiple lines in the spread sheet.
    """

    human_cohort = []
    chorts = list(df.loc[HUMAN_ARM_ID])
    for idx, val in enumerate(chorts[2:], start=2): 
        if not pd.isna(val):
            obj = {
               "arm_id": val,
               "arm_name": parse_clean_sv(df, HUMAN_ARM_NAME, idx),
               "study_population_description": parse_clean_sv(df, HUMAN_STUDY_POPULATION_DESCRIPTION, idx),
               "arm_type": check_spelling(parse_clean_sv(df, HUMAN_ARM_TYPE, idx),'arm_type'),
               "ethnicity": check_spelling(parse_clean_mv_split(df, HUMAN_ETHNICITY, idx),'ethnicity'),
               "race": check_spelling(parse_clean_mv_split(df, HUMAN_RACE, idx),'race'),
               "race_specify": parse_clean_mv_split(df, HUMAN_RACE_SPECIFY, idx),
               "description": parse_clean_sv(df, HUMAN_DESCRIPTION, idx),
               "sex_at_birth": check_spelling(parse_clean_mv_split(df, HUMAN_SEX_AT_BIRTH, idx),'sex_at_birth'),
               "age_event": parse_clean_sv(df, HUMAN_AGE_EVENT, idx),
               "subject_phenotype": parse_clean_sv(df, HUMAN_SUBJECT_PHENOTYPE,idx),
               "assessment_name": parse_clean_mv_split(df, HUMAN_ASSESSMENT_NAME, idx),
               "study_location": check_spelling(parse_clean_mv_split(df, HUMAN_STUDY_LOCATION, idx),'study_location'),
               "measured_behavioral_or_pyschological_factor": parse_clean_mv_split(df, HUMAN_MEASURED_BEHAVIORAL_OR_PYSCHOLOGICAL_FACTOR, idx),
               "measured_social_factor": parse_clean_mv_split(df, HUMAN_MEASURED_SOCIAL_FACTOR, idx),
               "sars_cov_2_symptoms": check_spelling(parse_clean_mv_split(df, HUMAN_SARS_COV_2_SYMPTOMS, idx),'sars_cov_2_symptoms'),
               "assessment_clinical_and_demographic_data_provenance": parse_clean_mv_split(df, HUMAN_ASSESSMENT_CLINICAL_AND_DEMOGRAPHIC_DATA_PROVENANCE, idx),
               "assessment_demographic_data_types_collected": parse_clean_mv_split(df, HUMAN_ASSESSMENT_DEMOGRAPHIC_DATA_TYPES_COLLECTED, idx),
               "sars_cov_2_history": check_spelling(parse_clean_mv_split(df, HUMAN_SARS_COV_2_HISTORY, idx),'sars_cov_2_history'),
               "sars_cov_2_vaccine_type": check_spelling(parse_clean_mv_split(df, HUMAN_SARS_COV_2_VACCINE_TYPE, idx),'sars_cov_2_vaccine_type-subject'),
               "covid_19_disease_severity": check_spelling(parse_clean_mv_split(df, HUMAN_COVID_19_DISEASE_SEVERITY, idx),'covid_19_disease_severity'),
               "post_covid_19_symptoms": parse_clean_mv_split(df, HUMAN_POST_COVID_19_SYMPTOMS, idx),
               "covid_19_complications": parse_clean_mv_split(df, HUMAN_COVID_19_COMPLICATIONS, idx)
            }
            human_cohort.append(obj)
        else:
            pass
    
    template['study_human_cohort'] = human_cohort


def parse_model_organism(df, template):
    """Parse the Subject Type: model organism

       Some rows in this section contain single value and some contain multi-valued content

       There may be serveral entries with one entry in each column
       spanning multiple lines in the spread sheet.
    """
    model_cohort = []
    chorts = list(df.loc[MODEL_ARM_ID])
    for idx, val in enumerate(chorts[2:], start=2): 
        if not pd.isna(val):
            obj = {
               "arm_id": val,
               "arm_name": parse_clean_sv(df, MODEL_ARM_NAME, idx),
               "study_population_description": parse_clean_sv(df, MODEL_STUDY_POPULATION_DESCRIPTION, idx),
               "arm_type": check_spelling(parse_clean_sv(df, MODEL_ARM_TYPE, idx),'arm_type'),
               "genus_and_species": check_spelling(parse_clean_mv_split(df, MODEL_SPECIES, idx),'genus_and_species'),
               "biosample_type": check_spelling(parse_clean_mv_split(df, MODEL_BIOSAMPLE_TYPE, idx),'biosample_type'),
               "strain_characteristics": parse_clean_mv_split(df, MODEL_STRAIN_CHARACTERISTICS, idx),
               "sex_at_birth": check_spelling(parse_clean_mv_split(df, MODEL_SEX_AT_BIRTH, idx),'sex_at_birth'),
               "age_event": parse_clean_sv(df, MODEL_AGE_EVENT, idx),
               "subject_phenotype": parse_clean_sv(df, MODEL_SUBJECT_PHENOTYPE, idx),
               "study_location": check_spelling(parse_clean_mv_split(df, MODEL_STUDY_LOCATION, idx),'study_location'),
               "sars_cov_2_history": check_spelling(parse_clean_mv_split(df, MODEL_SARS_COV_2_HISTORY, idx),'sars_cov_2_history'),
               "sars_cov_2_vaccine_type": check_spelling(parse_clean_mv_split(df, MODEL_SARS_COV_2_VACCINE_TYPE, idx),'sars_cov_2_vaccine_type-subject'),
               "covid_19_disease_severity": check_spelling(parse_clean_mv_split(df, MODEL_COVID_19_DISEASE_SEVERITY, idx),'covid_19_disease_severity'),
               "post_covid_19_symptoms": parse_clean_mv_split(df, MODEL_POST_COVID_19_SYMPTOMS, idx),
               "covid_19_complications": parse_clean_mv_split(df, MODEL_COVID_19_COMPLICATIONS, idx)
            }
            model_cohort.append(obj)
        else:
            pass
    
    template['study_model_cohort'] = model_cohort


def parse_planned_visit(df, template):
    """Parse the Planned Visit section

       There may be serveral entries with one entry in each column
       spanning multiple lines in the spread sheet.
    """

    plannedVisit = []
    visits = list(df.loc[VISIT_ID])
    for idx, val in enumerate(visits[2:], start=2):  
        if not pd.isna(val):
            obj = {
               "visit_id": val,
               "visit_name": parse_clean_sv(df, VISIT_NAME, idx),
               "visit_order_number": parse_sv(df, VISIT_ORDER_NUMBER, idx),
               "visit_min_start_day": parse_clean_sv(df, VISIT_MIN_START_DAY, idx),
               "visit_max_start_day": parse_clean_sv(df, VISIT_MAX_START_DAY, idx),
               "visit_start_rule": parse_clean_sv(df, VISIT_START_RULE, idx)
            }
            plannedVisit.append(obj)
        else:
            pass
    template['planned_visit'] = plannedVisit


def parse_experiment(df, template):
    """Parse the Experiment section

       Some rows in this section contain single value and some contain multi-valued content

       There may be serveral entries with one entry in each column
       spanning multiple lines in the spread sheet.
    """
    experiment = []
    experiments = list(df.loc[ASSOCIATED_ARM_IDS])
    for idx, val in enumerate(experiments[2:], start=2): 
        if not pd.isna(val):
            obj = {
               "arm_id": [cleanData(item) for item in re.split(" I ",val)],
               "associated_first_planned_visit_id": [cleanData(item) for item in re.split(" I ", df.loc[ASSOCIATED_FIRST_PLANNED_VISIT_ID][idx])],
               "assay_type": check_spelling(parse_clean_sv(df, ASSAY_TYPE, idx),'assay_type'),
               "experiment_name": parse_clean_sv(df, EXPERIMENT_NAME, idx),
               "experiment_results_file_name": parse_clean_sv(df, EXPERIMENT_RESULTS_FILE_NAME, idx),
               "biospecimen_type": check_spelling(parse_clean_mv_split(df, BIOSPECIMEN_TYPE, idx),'biospecimen_type'),
               "biospecimen_collection_point": check_spelling(parse_clean_mv_split(df, BIOSPECIMEN_COLLECTION_POINT, idx),'biospecimen_collection_point'),
               "sars_cov_2_antigen": check_spelling(parse_clean_mv_split(df, SARS_COV_2_ANTIGEN, idx),'sars_cov_2_antigen'),
               "assay_use": parse_clean_sv(df, ASSAY_USE, idx),
               "manufacture": parse_clean_sv(df, MANUFACTURER, idx),
               "catalog_number": parse_clean_sv(df, CATALOG_NUMBER, idx),
               "virus_target": list(set(check_spelling(parse_clean_mv_split(df, VIRUS_TARGET, idx),'virus_target'))),
               "antibody_isotype": parse_clean_mv_split(df, ANTIBODY_ISOTYPE, idx),
               "reporting_units": parse_clean_sv(df, REPORTING_UNITS, idx),
               "assay_reporting_format": parse_clean_sv(df, ASSAY_REPORTING_FORMAT, idx)
            }
            experiment.append(obj)
        else:
            pass
                                                                                 
    template['experiment'] = experiment

def parse_status_note(df, template):
    """Parse the Study Design section"""

    template['status_note'] = re.sub('[^a-zA-Z0-9 \n\.]',' ', replace_na(parse_clean_sv(df, STATUS_NOTE, 2)))

