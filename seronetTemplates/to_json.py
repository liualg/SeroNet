#!/usr/bin/env python
# coding: utf-8

'''
This script is compatibale with Registry Version v1.2.3 - 1.2.5
    - Please look at other template 
    - Added '*' to SARS-CoV-2 Antigen* (row 163, column B)
    - new assumptions from 1.2.5
    Assumptions:
        - specfiy over actaul fields 
        - loose query for v1.2.3 - 1.2.5
        - Assessments planned visit day 
        - Treatment 
'''
import json
import pandas as pd
import numpy as np
import os, shutil
import inspect
import datetime as dt
from sys import platform
from glob import glob
import sys

from tqdm import tqdm

import argparse 

from dataclasses import dataclass, field
from dataclasses_json import DataClassJsonMixin

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

## Importing Functions and Dataclass
import seronetDataclass as seroClass
import seronetFunctions as seroFxn

# import warnings

# warnings.simplefilter("ignore")

# if platform == "darwin":
#     os.system('clear')
# else:
#     os.system('cls')


# ## Inputs 
# parser = argparse.ArgumentParser(description='Process some integers.')
# parser.add_argument('pubmedIds', metavar='N', type=int, nargs='+',
#                     help='an integer for the accumulator')
# parser.add_argument('--sum', dest='accumulate', action='store_const',
#                     const=sum, default=max,
#                     help='sum the integers (default: find the max)')

# args = parser.parse_args()


# class EnhancedJSONEncoder(json.JSONEncoder):
#         def default(self, o):
#             if is_dataclass(o):
#                 return asdict(o)
#             return super().default(o)

#########################################
######### Taking in Inputs ##############
#########################################

def create_json(PMID):
# Registry
    sheet_name = 'SeroNet Registry Template'
    map_sheet = 'Registry Definitions'


    # finding correct Box Base
    if platform == "darwin":
        box_base = "~/Library/CloudStorage/Box-Box/SeroNet Public Data"
    else: 
        print("User has windows")
        box_base = os.path.join("Users",os.getlogin(), "Box")


    #File Paths
    BASE_DIR = seroFxn.get_box_dir(box_base, PMID)

    try:
        df_path = glob(os.path.join(BASE_DIR,'templated_data',f'PMID{PMID}*eviewed.xlsm'))[0]
    except:
        print("** NOT using reviewed files **")
        try:
            df_path = glob(os.path.join(BASE_DIR,'templated_data',f'PMID{PMID}*.xlsm'))[0]
        except FileNotFoundError:
            sys.exit("ERROR:: Incorrect Template format. Cannot Find File")
        
        
    # df_path = os.path.join(BASE_DIR,'templated_data', file + ".xlsm")
    print("file path: {}".format(df_path))
    # df_path = os.path.join(BASE_DIR,'templated_data', "PMID34431693_registry.xlsm")

    # dictionary info
    # registryToImmportDict_file = os.path.join("dictionary", "registryToBasic.csv")

    # registryToImmportDict = pd.read_csv(registryToImmportDict_file,
    #                                     header=None, 
    #                                     index_col=0, 
    #                                     squeeze=True).to_dict()


    # Automate output... 
    OUT_DIR = os.path.join(BASE_DIR, 'ImmPort_templates-DR46') 
    # OUT_DIR = './33184236_test/'
    JSON_template = f'PMID{PMID}_study.xlsx'


    # Make Dir if it does not exist
    try:
        os.mkdir(OUT_DIR)
        print(f'\nCreating output directory - {OUT_DIR}')
    except FileExistsError:
        print('\nWill not create directory - already exists')
        pass

    # import file
    book = load_workbook(df_path)
    registry = book[sheet_name]
    registry.delete_cols(1)

    print("\n~~ File Information ~~")

    # In[3]:


    # Class names
    class_names = ['study_pubmed', 'study', 'study_personnel', 'study_file',
                   'study_link', 'study_categorization', 'study_design',
                   'protocol', 'condition_or_disease', 'Intervention Agent',
                   'study_details', 'inclusion_exclusion',
                   'Subject Type: human', 'Subject Type: model organism', 'planned_visit',
                   'Experiments', 'Experiment Samples'
                  ]



    VARS_TO_CLEAN = ['', 'N/A', 'n/a', np.nan, None]
    clean_other = VARS_TO_CLEAN + ['Other']

    sp = seroFxn.get_sections(registry, class_names)
    sp.append(200)

    #########################################
    ######        Main Loop       ###########
    #########################################

    # Looping through each section in the Registy template
    for section_number in range(len(sp)-1):
        temp_wb = Workbook()
        temp_ws = temp_wb.active

        #making a temp workbook to store each section. This will be turned into df
        for i in registry.iter_rows(values_only = True,
                                    min_row = sp[section_number]+1,
                                    max_row = sp[section_number+1]-1):
            temp_ws.append(i)
            
        max_row = temp_ws.max_row
        max_col = temp_ws.max_column
        seroFxn.remove_excess(temp_ws)
        
        df = pd.DataFrame(temp_ws.values)
        sub_section = registry.cell(row=sp[section_number], column = 1).value.strip()
        if sub_section == 'study':
            df = seroFxn.edit_df(df)
            
            STUDY = seroClass.study(
                df['Study Identifier'][1],
                df['Study Name'][1],
                df['Publication Title'][1],
                df['Study Objective'][1],
                df['Study Description'][1],
                df['Primary Institution Name'][1]
                )
        
        elif sub_section == 'study_pubmed':
            df = seroFxn.edit_df(df)
            
            STUDY_PUBMED = seroClass.study_pubmed(
                # df['Pubmed ID'][1]
                PMID
            )
            
        elif sub_section == 'study_personnel':
            df = seroFxn.edit_df(df)
            try:
                STUDY_PERSONNEL = seroClass.study_personnel(
                    df['Personnel ID'],
                    df['Honorific'],
                    df['Last Name'],
                    df['First Name'],
                    df['Suffixes'],
                    df['Organization'],
                    df['ORCID ID'],
                    df['Email'],
                    df['SeroNet Title In Study'],
                    df['Role In Study'],
                    df['Site Name']
                )
            except:
                print("trying older format: Title in Study")
                STUDY_PERSONNEL = seroClass.study_personnel(
                    df['Personnel ID'],
                    df['Honorific'],
                    df['Last Name'],
                    df['First Name'],
                    df['Suffixes'],
                    df['Organization'],
                    df['ORCID ID'],
                    df['Email'],
                    df['Title In Study'],
                    df['Role In Study'],
                    df['Site Name']
                )


        elif sub_section == 'study_file':
            df = seroFxn.edit_df(df)

            STUDY_FILE = seroClass.study_file(
                df['Study File Name'],
                df['Study File Description'],
                df['Study File Type']  
            )
            
        
        elif sub_section == 'study_link':
            df = seroFxn.edit_df(df)

            STUDY_LINK = seroClass.study_link(
                df['Link Name'],
                df['Value']
            )
                 
            
        elif sub_section == 'study_categorization':
            df = seroFxn.edit_df(df)

            STUDY_CATEGORIZATION = seroClass.study_categorization(
                df['Research Focus*'][1],
                df['Study Type'][1],
                df['Keywords'][1],
            )
        
        elif sub_section == 'study_design':
            df = seroFxn.edit_df(df)

            if (df.shape != (2,0)): # checking size. There has to be a better way to do this

                STUDY_DESIGN = seroClass.study_design(
                    df['Clinical Study Design'],
                    df['in silico Model Type*']
                )
            else:
                study_des = study_design()

        elif sub_section == 'protocol':
            df = seroFxn.edit_df(df)

            PROTOCOLS = seroClass.protocols(
                df['Protocol ID'],
                df['Protocol File Name'],
                df['Protocol Name'],
                df['Protocol Description'],
                df['Protocol Type'],
            )
            
        elif sub_section == 'condition_or_disease':
            df = seroFxn.edit_df(df)

            COD = seroClass.condition_or_disease(
                list(df['Reported Health Condition* '])
            )
            
        elif sub_section == 'Intervention Agent':
            df = seroFxn.edit_df(df)

            INTERVENTION_AGENT = seroClass.Intervention_Agent(
                list(df['SARS-CoV-2 Vaccine Type*'])  ### IF THIS IS NOTHING, change to 'NA'
            )

        elif sub_section == 'study_details':
            df = seroFxn.edit_df(df)
            STUDY_DETAILS = seroClass.study_details(
                df['Clinical Outcome Measure'][1],
                df['Enrollment Start Date'][1],
                df['Enrollment End Date'][1],
                df['Number of Study Subjects'][1],
                df['Age Unit'][1],
                df['Minimum Age'][1],
                df['Maximum Age'][1]
                )

        elif sub_section == 'inclusion_exclusion':
            df = seroFxn.edit_df(df)

            INCLUSION_EXCLUSION = seroClass.inclusion_exclusion(
                df['Inclusion ID'],
                df['Inclusion Criterion'],
                df['Inclusion Criterion Category']  
            )
            
        elif sub_section == 'Subject Type: human':
            df = seroFxn.edit_df(df)
            tem_was_here = df

            try:
                SUBJECT_HUMAN = seroClass.subject_type_human(
                    df['Arm ID'],
                    df['Arm Name'],
                    df['Study Population Description'],
                    df['Arm Type'],
                    df['Ethnicity*'],
                    df['Race*'],
                    df['Race Specify'],
                    df['Description'],
                    df['Sex at Birth*'],
                    df['Age Event'],
                    df['Subject Phenotype'],
                    df['Study Location*'],
                    df['Assessment Name'],
                    df['Measured Behavioral or Psychological Factor*'],
                    df['Measured Social Factor*'],
                    df['SARS-CoV-2 Symptoms*'],
                    df['Assessment_Clinical  and Demographic Data Provenance*'],
                    df['Assessment_Demographic Data Types Collected'],
                    df['SARS-CoV2 History*'],
                    df['SARS-CoV-2 Vaccine Type*'],
                    df['COVID-19 Disease Severity*'],
                    df['Post COVID-19 Symptoms'],
                    df['COVID-19 Complications']
                    )
            except:
                print("** using older version of subject human **")
                SUBJECT_HUMAN = seroClass.subject_type_human(
                    df['Arm ID'],
                    df['Arm Name'],
                    df['Study Population Description'],
                    df['Arm Type'],
                    df['Ethnicity*'],
                    df['Race*'],
                    df['Race Specify'],
                    df['Description'],
                    df['Sex at Birth*'],
                    df['Age Event'],
                    df['Subject Phenotype'],
                    df['Study Location*'],
                    df['Assessment Name'],
                    df['Measured Behavioral or Psychological Factor*'],
                    df['Measured Social Factor*'],
                    df['SARS-CoV-2 Symptoms*'],
                    df['Assessment_Clinical  and Demographic Data Provenance'],
                    df['Assessment_Demographic Data Types Collected'],
                    df['SARS-CoV2 History*'],
                    df['SARS-CoV-2 Vaccine Type*'],
                    df['COVID-19 Disease Severity*'],
                    df['Post COVID-19 Symptoms'],
                    df['COVID-19 Complications']
                    )

        elif sub_section == 'Subject Type: model organism':
            df = seroFxn.edit_df(df)

            SUBJECT_ORGANISM = seroClass.subject_type_mode_organism(
                df['Arm ID'],
                df['Arm Name'],
                df['Study Population Description'],
                df['Arm Type'],
                df['Species'],
                df['Biosample Type'],
                df['Strain Characteristics'],
                df['Sex at Birth*'],
                df['Age Event'],
                df['Subject Phenotype'],
                df['Study Location*'],
                df['SARS-CoV2 History*'],
                df['SARS-CoV-2 Vaccine Type*'],
                df['COVID-19 Disease Severity*'],
                df['Post COVID-19 Symptoms'],
                df['COVID-19 Complications']
                )
        
        elif sub_section == 'planned_visit':
            df = seroFxn.edit_df(df)

            PLANNED_VISIT = seroClass.planned_visit(
                df['Visit ID'],
                df['Visit Name'],
                df['Visit Order Number'],
                df['Visit Min Start Day'],
                df['Visit Max Start Day'],
                df['Visit Start Rule']                                
        )
        
        elif sub_section == 'Experiment Samples':
            sys.exit("ERROR:: Incorrect Template format. Need to follow v1.2.3")
        #     FileExistsError
        #     df = seroFxn.edit_df(df)

        #     EXPERIMENT_SAMPLES = seroClass.study_experiment_samples(
        #         df['Associated Arm ID'],
        #         df['Associated Planned Visit ID'],
        #         df['Expt Sample User Defined ID'],
        #         df['Biospecimen Type*'],
        #         df['Biospecimen Collection Point*']                             
        # )
        
        elif sub_section == 'Experiments':
            df = seroFxn.edit_df(df)

            EXPERIMENTS = seroClass.experiments(
                df['Associated Arm ID(s)'],
                df['Associated First Planned Visit ID'],
                df['Assay Type'],
                df['Experiment Name'],
                df['Experiment Results File Name'],
                df['Biospecimen Type*'],
                df['Biospecimen Collection Point'], 
                df['SARS-CoV-2 Antigen*'],
                df['Assay Use'],
                df['Manufacturer'],
                df['Catalog #'],
                df['Virus Target'],
                df['Antibody Isotype*'],
                df['Reporting Units'],
                df['Assay Reporting Format']
        )
            # print(INCLUSION_EXCLUSION)
            
            ImmPortStudy = seroClass.studySearch(
                Study_Identifier = STUDY.Study_Identifier,
                Study_Name = STUDY.Study_Name,
                Publication_Title = STUDY.Publication_Title,
                Study_Objective = STUDY.Study_Objective,
                Study_Description = STUDY.Study_Description,
                Primary_Institution_Name = STUDY.Primary_Institution_Name,
                Geriatric_subjects = INCLUSION_EXCLUSION.Geriatric_subjects,
                Pediatric_subjects = INCLUSION_EXCLUSION.Pediatric_subjects,
                Pregnant_subjects = INCLUSION_EXCLUSION.Pregnant_subjects,
                SARS_CoV_2_Antibodies_Measured = INCLUSION_EXCLUSION.SARS_CoV_2_Antibodies_Measured,
                Performance_metrics_included = INCLUSION_EXCLUSION.Performance_metrics_included,
                Survey_instrument_shared = INCLUSION_EXCLUSION.Survey_instrument_shared,
                WHO_disease_severity_scale_used = INCLUSION_EXCLUSION.WHO_disease_severity_scale_used

                )



            TEMPLATE = seroClass.Template(
                ImmPortStudy,
                STUDY_PUBMED,
                STUDY_PERSONNEL,
                STUDY_FILE,
                STUDY_LINK,
                STUDY_CATEGORIZATION,
                STUDY_DESIGN,
                PROTOCOLS,
                COD,
                INTERVENTION_AGENT,
                STUDY_DETAILS,
                INCLUSION_EXCLUSION,
                SUBJECT_HUMAN,
                SUBJECT_ORGANISM,
                PLANNED_VISIT,
                EXPERIMENTS
                )

            output_file = os.path.join(OUT_DIR,f"PMID{PMID}_json.json")

            # print(json.dumps(template, indent=4))
            # print(TEMPLATE.to_json(indent=4))
            # print(json.dumps(TEMPLATE))
            f = open(output_file, "w")
            print(TEMPLATE.to_json(indent=4), file = f)
            # print(json.dumps(template, indent=4), file = f)
            f.close()
            
            
        else:
            print(sub_section, ': does not exist')



    #######################################################
    #####    POST: copy log + other files to folder   #####
    #######################################################


    # for filename in glob(os.path.join(CD,BASE_DIR,"submitted_data", '*.*')):
    #     shutil.copy(filename, OUT_DIR)

    # print("\n")
    # for i in glob(os.path.join(OUT_DIR,"*.txt")):
    #     print(os.path.basename(i))
    # os.remove(PATH_pmid_basic_stdy_template)



# for unqiueID in args.pubmedIds:
#     create_full(str(unqiueID))
