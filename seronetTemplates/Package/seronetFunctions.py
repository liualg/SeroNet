import pandas as pd 
import numpy as np
import os, shutil
import inspect

from tqdm import tqdm

from dataclasses import dataclass, field
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# find file dir to save files base off PMID
def check_input(pmid):
    if len(pmid) == 8:
        return pmid
    
    else:
        try:
            return check_PMID(pmid.split("_")[1])
        except:
            print('PMID format invalid')
            return False

    
def get_box_dir(box_dir, pmid):
    box_base = os.path.abspath(os.path.expanduser(os.path.expandvars(box_dir)))
    depth = 3

    for dirpath, dirnames, filenames in os.walk(box_base):
        if dirpath[len(box_base):].count(os.sep) < depth:
            if "PMID_"+pmid in dirnames:
                return(os.path.join(dirpath,"PMID_"+pmid))


# Return start index for each section in the SeroNet Registry Template 
def get_sections(ws, class_names):
    temp = []
    for i in range(1,ws.max_row):
        if ws["A"][i].value in class_names:
            temp.append(i+1)
    return temp

def column_header_index(sheet):
    count = 0
    row_index = 1

    for row in sheet.iter_rows():
        if [cell.value for cell in row] == [None]*sheet.max_column:
            break
        row_index += 1
    
    return row_index

def row_header_index(sheet):
    count = 0
    col_index = 1

    for col in sheet.iter_cols():
        if [cell.value for cell in col] == [None]*sheet.max_row:
            break
        col_index += 1
    
    return col_index

def remove_excess(sheet):
    sheet.delete_rows(column_header_index(sheet),sheet.max_row)
    sheet.delete_cols(row_header_index(sheet),sheet.max_column)
    return sheet

def edit_df(df_temp):
    df_temp = df_temp.transpose()
    df_temp.columns = df_temp.iloc[0]
    return(df_temp.drop(df_temp.index[0]))


#################  Adding sections back to dataframes  ##################

def add_df(ws, input_df_class, add_header = True, stagger = 0):
    
    if isinstance(input_df_class, pd.DataFrame): #checking to see if it is a class
        df = input_df_class

    else: #if its a dataframe - just overwrite df to be the dataframe
                
        if input_df_class.ImmPortNAME == 'study_pubmed':
            df = pd.DataFrame({**vars(input_df_class)}, index=[0]).drop(['ImmPortNAME'], axis = 1)
        else:
            df = pd.DataFrame({**vars(input_df_class)}).drop(['ImmPortNAME'], axis = 1)
        
    df.columns = [i.replace("_"," ") for i in df.columns]

    if add_header == True:
        ws.append([])
        ws.append([input_df_class.ImmPortNAME])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(['']*stagger + r)
    else:
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(['']*stagger + r)
        









    