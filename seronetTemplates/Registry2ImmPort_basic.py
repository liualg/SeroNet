#!/usr/bin/env python
# coding: utf-8

# This script is compatibale with Registry Version v.1.1.0

import pandas as pd
import numpy as np
import os, shutil
import inspect
import datetime as dt
from sys import platform

from tqdm import tqdm

import argparse 

from dataclasses import dataclass, field
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

## Importing Functions and Dataclass
import seronetDataclass_v1_1_0 as seroClass
import seronetFunctions as seroFxn

import warnings

warnings.simplefilter("ignore")

try:
    os.system('clear')
except:
    os.system('cls')


#########################################
######### Taking in Inputs ##############
#########################################

def create_basic(PMID):
# Registry
    sheet_name = 'SeroNet Registry Template'
    map_sheet = 'Registry Definitions'


    # finding correct Box Base
    if platform == "darwin":
        box_base = "~/Library/CloudStorage/Box-Box/SeroNet Public Data"
    else: 
        print("User has windows")
        box_base = os.path.join("Users",os.getlogin(), "Box")


    #File Paths
    BASE_DIR = seroFxn.get_box_dir(box_base, PMID)


    try: 
        file = f"PMID{PMID}_v1.1.0"
    except:
        file = f"PMID{PMID}_Registry"
        
        
    df_path = os.path.join(BASE_DIR,'templated_data', file + ".xlsm")
    print("file path: {}".format(df_path))
    # df_path = os.path.join(BASE_DIR,'templated_data', "PMID34431693_registry.xlsm")

    # dictionary info
    registryToImmportDict_file = os.path.join("dictionary", "registryToBasic.csv")

    registryToImmportDict = pd.read_csv(registryToImmportDict_file,
                                        header=None, 
                                        index_col=0, 
                                        squeeze=True).to_dict()

    # ImmPort Templates (link to web?)
    PATH_basic_stdy_template = os.path.join("template", "basic_study_design.xlsx")
    PATH_protocols = os.path.join("template", "protocols.xlsx")



    # Automate output... 
    OUT_DIR = os.path.join(BASE_DIR, 'ImmPort_templates') 
    # OUT_DIR = './33184236_test/'
    PATH_pmid_basic_stdy_template = f'PMID{PMID}_study.xlsx'

    BASIC_STUDY_TEMPLATE = f'PMID{PMID}_basic'
    PROTOCOL_TEMPLATE = f'PMID{PMID}_protocol'


    # Make Dir if it does not exist
    try:
        os.mkdir(OUT_DIR)
        print(f'Creating output directory - {OUT_DIR}')
    except FileExistsError:
        print('Will not create directory - already exists')
        pass

    # import file
    book = load_workbook(df_path)
    registry = book[sheet_name]
    registry.delete_cols(1)

    print("\n~~ File Information ~~")

    # In[3]:


    # Class names
    class_names = ['study_pubmed', 'study', 'study_personnel', 'study_file',
                   'study_link', 'study_categorization', 'study_design',
                   'protocol', 'condition_or_disease', 'Intervention Agent',
                   'study_details', 'inclusion_exclusion',
                   'Subject Type: human', 'Subject Type: model organism', 'planned_visit',
                   'Study Experiments', 'Study Experiment Samples', 'Reagent', 
                   'Results for Serology Assays'
                  ]


    VARS_TO_CLEAN = ['', 'N/A', 'n/a', np.nan, None]


    # In[4]:


    sp = seroFxn.get_sections(registry, class_names)

    #########################################
    ######        Main Loop       ###########
    #########################################

    # Looping through each section in the Registy template
    for section_number in range(len(sp)-1):
        temp_wb = Workbook()
        temp_ws = temp_wb.active

        #making a temp workbook to store each section. This will be turned into df
        for i in registry.iter_rows(values_only = True,
                                    min_row = sp[section_number]+1,
                                    max_row = sp[section_number+1]-1):
            temp_ws.append(i)
            
        max_row = temp_ws.max_row
        max_col = temp_ws.max_column
        seroFxn.remove_excess(temp_ws)
        
        df = pd.DataFrame(temp_ws.values)
        sub_section = registry.cell(row=sp[section_number], column = 1).value.strip()
        
        if sub_section == 'study':
            df = seroFxn.edit_df(df)
            
            STUDY = seroClass.study(
                df['Study Identifier'][1],
                df['Study Name'][1],
                df['Publication Title'][1],
                df['Study Objective'][1],
                df['Study Description'][1],
                df['Primary Institution Name'][1]
                )
        
        elif sub_section == 'study_pubmed':
            df = seroFxn.edit_df(df)
            
            STUDY_PUBMED = seroClass.study_pubmed(
                # df['Pubmed ID'][1]
                PMID
            )
            
        elif sub_section == 'study_personnel':
            df = seroFxn.edit_df(df)

            STUDY_PERSONNEL = seroClass.study_personnel(
                df['Personnel ID'],
                df['Honorific'],
                df['Last Name'],
                df['First Name'],
                df['Suffixes'],
                df['Organization'],
                df['ORCID ID'],
                df['Email'],
                df['Title In Study'],
                df['Role In Study'],
                df['Site Name']
            )


        elif sub_section == 'study_file':
            df = seroFxn.edit_df(df)

            STUDY_FILE = seroClass.study_file(
                df['Study File Name'],
                df['Study File Description'],
                df['Study File Type']  
            )
            
        
        elif sub_section == 'study_link':
            df = seroFxn.edit_df(df)

            STUDY_LINK = seroClass.study_link(
                df['Link Name'],
                df['Value']
            )
                 
            
        elif sub_section == 'study_categorization':
            df = seroFxn.edit_df(df)

            STUDY_CATEGORIZATION = seroClass.study_categorization(
                df['Research Focus*'][1],
                df['Study Type'][1],
                df['Keywords'][1],
            )
        
        elif sub_section == 'study_design':
            df = seroFxn.edit_df(df)

            if (df.shape != (2,0)): # checking size. There has to be a better way to do this

                STUDY_DESIGN = seroClass.study_design(
                    df['Clinical Study Design'],
                    df['in silico Model Type*']
                )
            else:
                study_des = study_design()

        elif sub_section == 'protocol':
            df = seroFxn.edit_df(df)

            PROTOCOLS = seroClass.protocols(
                df['Protocol ID'],
                df['Protocol File Name'],
                df['Protocol Name'],
                df['Protocol Description'],
                df['Protocol Type'],
            )
            
        elif sub_section == 'condition_or_disease':
            df = seroFxn.edit_df(df)

            COD = seroClass.condition_or_disease(
                list(df['Reported Health Condition* '])
            )
            
        elif sub_section == 'Intervention Agent':
            df = seroFxn.edit_df(df)

            INTERVENTION_AGENT = seroClass.Intervention_Agent(
                list(df['SARS-CoV-2 Vaccine Type*'])  ### IF THIS IS NOTHING, change to 'NA'
            )

        elif sub_section == 'study_details':
            df = seroFxn.edit_df(df)
            STUDY_DETAILS = seroClass.study_details(
                df['Clinical Outcome Measure'][1],
                df['Enrollment Start Date'][1],
                df['Enrollment End Date'][1],
                df['Number of Study Subjects'][1],
                df['Age Unit'][1],
                df['Minimum Age'][1],
                df['Maximum Age'][1]
                )

        elif sub_section == 'inclusion_exclusion':
            df = seroFxn.edit_df(df)

            INCLUSION_EXCLUSION = seroClass.inclusion_exclusion(
                df['Inclusion ID'],
                df['Inclusion Criterion'],
                df['Inclusion Criterion Category']  
            )
            
        elif sub_section == 'Subject Type: human':
            df = seroFxn.edit_df(df)
            tem_was_here = df

            SUBJECT_HUMAN = seroClass.subject_type_human(
                df['Arm ID'],
                df['Arm Name'],
                df['Study Population Description'],
                df['Arm Type'],
                df['Ethnicity*'],
                df['Race*'],
                df['Race Specify'],
                df['Description'],
                df['Sex at Birth*'],
                df['Age Event'],
                df['Subject Phenotype'],
                df['Study Location*'],
                df['Assessment Name'],
                df['Measured Behavioral or Psychological Factor*'],
                df['Measured Social Factor*'],
                df['SARS-CoV-2 Symptoms*'],
                df['Assessment_Clinical  and Demographic Data Provenance'],
                df['Assessment_Demographic Data Types Collected'],
                df['SARS-CoV2 History*'],
                df['SARS-CoV-2 Vaccine Type*'],
                df['COVID-19 Disease Severity*'],
                df['Post COVID-19 Symptoms'],
                df['COVID-19 Complications']
                )

        elif sub_section == 'Subject Type: model organism':
            df = seroFxn.edit_df(df)

            SUBJECT_ORGANISM = seroClass.subject_type_mode_organism(
                df['Arm ID'],
                df['Arm Name'],
                df['Study Population Description'],
                df['Arm Type'],
                df['Species'],
                df['Biosample Type'],
                df['Strain Characteristics'],
                df['Sex at Birth*'],
                df['Age Event'],
                df['Subject Phenotype'],
                df['Study Location*'],
                df['SARS-CoV2 History*'],
                df['SARS-CoV-2 Vaccine Type*'],
                df['COVID-19 Disease Severity*'],
                df['Post COVID-19 Symptoms'],
                df['COVID-19 Complications']
                )
        
        elif sub_section == 'planned_visit':
            df = seroFxn.edit_df(df)

            PLANNED_VISIT = seroClass.planned_visit(
                df['Visit ID'],
                df['Visit Name'],
                df['Visit Order Number'],
                df['Visit Min Start Day'],
                df['Visit Max Start Day'],
                df['Visit Start Rule']                                
        )
        
      
        else:
            print(sub_section, ':: Will not be used')



    # # Sending back to ImmPort Template #



    ImmPortClassNames = ['study', 'study_categorization', 'study_2_condition_or_disease',
                   'arm_or_cohort','study_personnel', 'planned_visit', 'inclusion_exclusion',
                   'study_2_protocol', 'study_file', 'study_link', 'study_pubmed'] 


    # In[9]:


    shutil.copy(PATH_basic_stdy_template, PATH_pmid_basic_stdy_template)
    basic_stdy_template = load_workbook(PATH_pmid_basic_stdy_template)
    bst_ws = basic_stdy_template['basic_study_design.txt']

    se = seroFxn.get_sections(bst_ws, ImmPortClassNames)

    #########################################
    ######### Basic Study Template ##########
    #########################################
    clean_vaccine = VARS_TO_CLEAN + ['Other']

    if len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type) + len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
        
        vaccine_name, vaccine_type = seroFxn.get_vaccine(set(list(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type) + list(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type)),
                        clean_vaccine)
        
        
    elif len(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type):
        
        vaccine_name, vaccine_type = seroFxn.get_vaccine(set(list(SUBJECT_HUMAN.SARS_CoV_2_Vaccine_Type)),
                        clean_vaccine)
                                
    elif len(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type):
        
        vaccine_name, vaccine_type = seroFxn.get_vaccine(set(list(SUBJECT_ORGANISM.SARS_CoV_2_Vaccine_Type)),
                        clean_vaccine)
        
    else:
        
        vaccine_name = []



    if SUBJECT_HUMAN and SUBJECT_ORGANISM:
        print('Both human and model organism are used')
        usr_id = list(SUBJECT_HUMAN.User_Defined_ID) + list(SUBJECT_ORGANISM.User_Defined_ID)
        name = list(SUBJECT_HUMAN.Name) + list(SUBJECT_ORGANISM.Name)
        description = list(SUBJECT_HUMAN.Description) + list(SUBJECT_ORGANISM.Description)
        type_report = list(SUBJECT_HUMAN.Type_Reported) + list(SUBJECT_ORGANISM.Type_Reported)
        
        if set(SUBJECT_HUMAN.User_Defined_ID).intersection(SUBJECT_ORGANISM.User_Defined_ID):
            print("\n*** ERROR: Cannot have same User Defined ID's for AOCs:")
            print(set(SUBJECT_HUMAN.User_Defined_ID).intersection(SUBJECT_ORGANISM.User_Defined_ID))
            
    if SUBJECT_HUMAN and not SUBJECT_ORGANISM:
        print('Only human is used')
        usr_id = list(SUBJECT_HUMAN.User_Defined_ID)
        name = list(SUBJECT_HUMAN.Name)
        description = list(SUBJECT_HUMAN.Description)
        type_report = list(SUBJECT_HUMAN.Type_Reported)
        
    if SUBJECT_ORGANISM and not SUBJECT_HUMAN:
        print('Only organism is used')
        usr_id = list(SUBJECT_ORGANISM.User_Defined_ID)
        name = list(SUBJECT_ORGANISM.Name)
        description = list(SUBJECT_ORGANISM.Description)
        type_report = list(SUBJECT_ORGANISM.Type_Reported)



    temp_wb = Workbook()
    temp_ws = temp_wb.active

    #making a temp workbook to store each section. This will be turned into df

    # This is the first part of the df
    for i in bst_ws.iter_rows(values_only = True,
                              max_row = se[2]+1):
        temp_ws.append(i)
        
    # Starting at Arm or Cohort  

    AOC = seroClass.arm_or_cohort(
        list(filter(None, usr_id)),
        list(filter(None, name)),
        list(filter(None, description)),
        list(filter(None, type_report))
    )

    # temp_ws.append(['User Defined ID', 'Name','Description','Type Reported'])
    seroFxn.add_df(temp_ws, AOC)

    seroFxn.add_df(temp_ws, STUDY_PERSONNEL)
    seroFxn.add_df(temp_ws, PLANNED_VISIT)
    seroFxn.add_df(temp_ws, INCLUSION_EXCLUSION)

    # # Add protocol
    temp_ws.append([])
    temp_ws.append([PROTOCOLS.ImmPortNAME])
    temp_ws.append(['Protocol ID', PROTOCOLS.Protocol_Name[1]])

    seroFxn.add_df(temp_ws, STUDY_FILE)
    seroFxn.add_df(temp_ws, STUDY_LINK)

    seroFxn.add_df(temp_ws, STUDY_PUBMED)


    # #### Filling in Study Info Sections 



    registryToImmportDict = pd.read_csv(registryToImmportDict_file, 
                                        header=None, 
                                        index_col=0, 
                                        squeeze=True).to_dict()

    # registryToImmportDict
    registryDict = {**vars(STUDY),
                    **vars(STUDY_CATEGORIZATION),
                    **vars(STUDY_DESIGN),
                    **vars(STUDY_DETAILS),
                    **vars(STUDY_PUBMED),
                    **vars(COD)}

    registryDict['SARS-CoV-2_Vaccine_Type'] = ['N/A']

    if len([x for x in vaccine_name if x not in VARS_TO_CLEAN]) != 0:
            registryDict['SARS-CoV-2_Vaccine_Type'] = [x for x in vaccine_name if x not in VARS_TO_CLEAN]


    # Looping through ImmPort Template to get the correct order of the 'study' section
    for se_number in range(se[0],se[3]):    
     
        if temp_ws["A"][se_number].value != None and registryToImmportDict.get(temp_ws["A"][se_number].value) != None:
            
            # Using a mapping key + using info in our classes to map the data
            reg_key = registryToImmportDict.get(temp_ws["A"][se_number].value).strip().replace(' ',"_").replace('*',"")
            try:
                # If input is a list, we will turn it into a string (since it cant be a list)
                if type(registryDict.get(reg_key)) == list:
                    temp_ws["B"][se_number].value = ', '.join(registryDict.get(reg_key))
                else:
                    temp_ws["B"][se_number].value = registryDict.get(reg_key)

            except:
                print(f"{reg_key} did not work")
                


    bsd = pd.DataFrame(temp_ws.values).replace({None: '', 'None': ''})
    # bsd.to_excel(os.path.join(OUT_DIR, f'PMID{PMID}_study.xlsx'), index=False, header = False)
    bsd.to_csv(os.path.join(OUT_DIR,f'{BASIC_STUDY_TEMPLATE}.txt'),
               header = False,
               index = False,
               sep = '\t')


    #########################################
    ############## Protocol #################
    #########################################

    #Load Protocol
    protocol_ws = load_workbook(PATH_protocols)['protocols.txt']
    protocol_ws = seroFxn.remove_excess(protocol_ws)
    seroFxn.add_df(protocol_ws, PROTOCOLS, add_header = False, stagger = 1)



    protocol_df = pd.DataFrame(protocol_ws.values).replace({None: '', 'None': ''})
    protocol_df.to_csv(os.path.join(OUT_DIR,f'{PROTOCOL_TEMPLATE}.txt'),
                       header = False, 
                       index = False,
                       sep = '\t')
    try:
        shutil.copy(os.path.join(BASE_DIR,'templated_data', f'{PROTOCOLS.Protocol_Name[1]}.txt'),
                    os.path.join(OUT_DIR,f'{PROTOCOLS.Protocol_Name[1]}.txt'))
    except:
        print(f'File: {PROTOCOLS.Protocol_Name[1]}.txt does not exist')



    #########################################
    #####    POST: copy log to folder   #####
    #########################################

    CD = os.getcwd()
    today = dt.datetime.today().strftime('%Y_%m_%d')

    try:
        os.mkdir(os.path.join(CD,BASE_DIR,"log"))
        print(f'Creating log')
    except FileExistsError:
        print('Will not create log - already exists')
        pass

    shutil.copyfile(os.path.join(CD,"log",f"Registry_{today}.log"), os.path.join(CD,BASE_DIR,"log",f"Registry_{today}.log"))
    os.remove(PATH_pmid_basic_stdy_template)
